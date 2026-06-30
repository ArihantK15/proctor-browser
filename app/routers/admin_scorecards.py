"""Scorecard and export router. Extracted from admin.py."""

import asyncio
import io
import json
import logging
import os
import zipfile
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Request, HTTPException, Body
from fastapi.responses import StreamingResponse

from ..auth import require_admin
from ..auth.scope import resolve_scope, scope_to_teacher_ids, assert_session_accessible, apply_teacher_scope
from ..database import async_table as _atable
from ..repositories.sessions import (
    assert_session_owned as _assert_session_owned,
    cohort_roll_numbers as _cohort_roll_numbers,
    fetch_all_results as _fetch_all_results,
    stream_csv_results as _stream_csv_results,
)
from ..repositories.questions import load_questions as _load_questions, load_exam_config as _load_exam_config
from ..services.risk import compute_risk_score, _is_violation
from ..utils import _safe_filename, _html_escape, _xlsx_safe, fmt_ist, now_ist
from ..models import SessionStatus, RESULT_STATUSES
from .. import cache as _cache
from ..services.sessions import collect_session_screenshots as _collect_session_screenshots, match_screenshot_for_violation as _match_screenshot_for_violation, match_room_screenshot_for_violation as _match_room_screenshot_for_violation, match_context_screenshots_for_violation as _match_context_screenshots_for_violation
from ..limiter import limiter
from ..models import EmailScorecardsIn
from ..services.scorecard import _build_scorecard_pdf, resolve_student_name
from ..jobs import enqueue_job, send_scorecard_email_job

_admin_log = logging.getLogger("admin")
logger = logging.getLogger(__name__)

router = APIRouter(prefix="")


@router.get("/api/v1/export-csv")
@limiter.limit("10/minute")
async def export_csv(request: Request, exam_id: Optional[str] = None, group_id: Optional[str] = None, batch: Optional[str] = None):
    teacher = await require_admin(request)
    scope = await resolve_scope(teacher, request)
    tids = await scope_to_teacher_ids(scope)
    roll_numbers = await _cohort_roll_numbers(tids, group_id=group_id, batch=batch)
    return StreamingResponse(
        _stream_csv_results(teacher["id"], exam_id=exam_id, max_rows=5000, teacher_ids=tids,
                            roll_numbers=roll_numbers),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=results.csv"})


@router.get("/api/v1/export-excel")
@limiter.limit("10/minute")
async def export_excel(request: Request, exam_id: Optional[str] = None, group_id: Optional[str] = None, batch: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell import WriteOnlyCell

    teacher = await require_admin(request)
    scope = await resolve_scope(teacher, request)
    tids = await scope_to_teacher_ids(scope)
    roll_numbers = await _cohort_roll_numbers(tids, group_id=group_id, batch=batch)
    results = await _fetch_all_results(teacher["id"], exam_id=exam_id, teacher_ids=tids,
                                       roll_numbers=roll_numbers)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    safe_eid = "".join(c for c in (exam_id or "all") if c.isalnum() or c in "-_")[:24]
    ws.title = f"Results_{safe_eid}" if safe_eid else "Results"

    headers = ["Timestamp", "Session ID", "Roll Number", "Full Name",
               "Email", "Score", "Total", "Percentage", "Time (min)",
               "Violations", "Risk Score", "Risk Label"]

    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    styled_headers = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align
        styled_headers.append(c)
    ws.append(styled_headers)

    risk_fills = {
        "Low":    PatternFill("solid", fgColor="D1FAE5"),
        "Medium": PatternFill("solid", fgColor="FEF3C7"),
        "High":   PatternFill("solid", fgColor="FEE2E2"),
    }

    for s in results:
        try:
            pct_val = float(s.get("percentage") or 0)
        except Exception:
            logger.debug("excel: bad percentage value %r for %s",
                         s.get("percentage"), s.get("session_key"))
            pct_val = 0.0
        try:
            secs = int(s.get("time_taken_secs") or 0)
        except Exception:
            logger.debug("excel: bad time_taken_secs %r for %s",
                         s.get("time_taken_secs"), s.get("session_key"))
            secs = 0
        mins = round(secs / 60, 2) if secs else 0

        row = [
            _xlsx_safe(s.get("submitted_at", "")),
            _xlsx_safe(s.get("session_id", "")),
            _xlsx_safe(s.get("roll_number", "")),
            _xlsx_safe(s.get("full_name", "")),
            _xlsx_safe(s.get("email", "")),
            s.get("score", 0),
            s.get("total", 0),
            pct_val,
            mins,
            s.get("violation_count", 0),
            s.get("risk_score", ""),
            _xlsx_safe(s.get("risk_label", "")),
        ]
        label = s.get("risk_label")
        fill = risk_fills.get(str(label) if label else "")
        if fill:
            cell = WriteOnlyCell(ws, value=row[-1])
            cell.fill = fill
            row[-1] = cell
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"results_{safe_eid or 'all'}_{now_ist().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── PDF helper: confidence parsing ────────────────────────────────

_CONF_MAP = {
    "face_missing": 0.95, "multiple_faces": 0.92,
    "wrong_person": 0.78, "eyes_closed": 0.88,
    "earphone_detected": 0.72, "cheat_object_detected": 0.85,
    "gaze_away": 0.70, "head_away": 0.80,
    "voice_detected": 0.75, "window_focus_lost": 0.99,
    "tab_hidden": 0.99, "lighting_issue": 0.90,
    "shortcut_blocked": 0.99, "camera_covered": 0.95,
}


def _pdf_get_conf(vtype: str, details) -> str:
    det = str(details or "")
    for prefix in ("confidence:", "conf:"):
        if prefix in det:
            try:
                raw = det.split(f"{prefix}")[1].split("|")[0].strip() if prefix == "confidence:" else det.split(f"{prefix}")[1].split(" ")[0].strip()
                if prefix == "conf:":
                    val = float(raw)
                    return f"{int(val)}%" if val > 1 else f"{int(val*100)}%"
                return raw if "%" in raw else f"{raw}%"
            except Exception:
                logger.debug("_pdf_get_conf: bad %s format in %s", prefix, vtype)
    return f"{int(_CONF_MAP.get(vtype, 0.75) * 100)}%"


def _pdf_clean_details(details) -> str:
    det = str(details or "")
    raw = det.split("| confidence:")[0].strip()[:40] if "| confidence:" in det else det[:40]
    return _html_escape(raw)


async def _pdf_fetch_violations(session_id: str, tid: str) -> list[dict[str, Any]]:
    result = await _atable("violations").select("*")\
        .eq("session_key", session_id)\
        .eq("teacher_id", str(tid))\
        .order("created_at").execute()
    # high/medium severity AND an actual student violation — exclude high-severity
    # DIAGNOSTIC events (model_load_failed, proctor_failed, …) which the severity
    # filter alone would let into the report's Violations table.
    return [v for v in (result.data or [])
            if v["severity"] in ("high", "medium") and _is_violation(v["violation_type"])]


async def _pdf_fetch_answers(session_id: str, tid: str) -> list[dict[str, Any]]:
    result = await _atable("answers").select("*")\
        .eq("session_key", session_id)\
        .eq("teacher_id", str(tid)).execute()
    return result.data or []


def _pdf_build_info_table(exam: dict[str, Any], raw_violations: list[Any], risk: dict[str, Any]):
    from reportlab.lib import colors as _c
    from reportlab.platypus import Table, TableStyle
    _score = exam.get("score", 0) or 0
    _total = exam.get("total", 0) or 0
    _pct = exam.get("percentage")
    if _pct in (None, 0) and _total:
        _pct = round(_score / _total * 100, 1)
    _pct = _pct or 0
    info = [
        ["Field", "Value"],
        ["Full Name", exam["full_name"]],
        ["Roll Number", exam["roll_number"]],
        ["Email", exam.get("email", "")],
        ["Submitted At", fmt_ist(exam.get("submitted_at", ""))],
        ["Score", f"{_score}/{_total} ({_pct}%)"],
        ["Time Taken", f"{exam.get('time_taken_secs', 0)}s ({exam.get('time_taken_secs', 0)//60}m {exam.get('time_taken_secs', 0)%60}s)"],
        ["Total Violations", str(len(raw_violations))],
        ["Behavioral Risk Score", f"{risk['risk_score']}/100 \u2014 {risk['label']}"],
    ]
    t = Table(info, colWidths=[160, 310])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _c.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), _c.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_c.HexColor("#f0f4ff"), _c.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, _c.grey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    return t


def _pdf_build_violations_table(raw_violations: list[Any]):
    from reportlab.lib import colors as _c
    from reportlab.platypus import Table, TableStyle

    total_conf_vals = []
    vd = [["#", "Type", "Severity", "Time", "Conf", "Details"]]
    for i, v in enumerate(raw_violations, 1):
        conf_str = _pdf_get_conf(v["violation_type"], v.get("details"))
        try:
            total_conf_vals.append(float(conf_str.strip("%")) / 100)
        except Exception:
            logger.debug("scorecards: confidence float parse failed", exc_info=True)
        ts_part = ""
        if v.get("created_at"):
            _fmted = fmt_ist(v["created_at"])
            _comma = _fmted.find(",")
            ts_part = _fmted[_comma + 1:].replace("IST", "").strip() if _comma >= 0 else _fmted
        vd.append([
            str(i),
            v["violation_type"].replace("_", " ").title()[:22],
            v["severity"].upper(),
            ts_part, conf_str,
            _pdf_clean_details(v.get("details"))[:35],
        ])

    vt = Table(vd, colWidths=[20, 120, 55, 70, 40, 165])
    vt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _c.HexColor("#c0392b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), _c.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_c.HexColor("#fff5f5"), _c.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, _c.grey),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
    ]))
    return vt, total_conf_vals


def _pdf_build_confidence_summary(total_conf_vals: list[Any], raw_count: int):
    from reportlab.lib import colors as _c
    from reportlab.platypus import Table, TableStyle
    if not total_conf_vals:
        return None
    avg_conf = sum(total_conf_vals) / len(total_conf_vals)
    high_conf = len([c for c in total_conf_vals if c >= 0.85])
    reliability = "High" if avg_conf >= 0.85 else "Medium" if avg_conf >= 0.70 else "Low"
    ct = Table(
        [[f"Overall Detection Confidence: {avg_conf:.0%}",
          f"High Confidence Violations: {high_conf}/{raw_count}",
          f"Reliability: {reliability}"]],
        colWidths=[160, 160, 150])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _c.HexColor("#eafaf1")),
        ("TEXTCOLOR", (0, 0), (-1, -1), _c.HexColor("#1e8449")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, _c.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    return ct


def _pdf_find_evidence(session_id: str, exam: dict[str, Any], raw_violations: list[Any], tid: str) -> list[Any]:
    roll = exam.get("roll_number") or (
        session_id.rsplit("_", 1)[0] if "_" in session_id else session_id[:20])
    paths = _collect_session_screenshots(roll, tid)
    items = []
    for idx, v in enumerate(raw_violations, 1):
        match = _match_screenshot_for_violation(v, paths)
        if match is not None and match.exists():
            # Phone-cam companion captured at the same instant (None if the
            # student didn't pair a phone) — rendered side by side.
            room = _match_room_screenshot_for_violation(v, paths)
            room = room if (room is not None and room.exists()) else None
            # Pre-violation context frames (t-3s..t-0) for appeal-critical
            # flags — the lead-up, rendered as a small strip beneath the flag.
            context = [p for p in _match_context_screenshots_for_violation(v, paths)
                       if p is not None and p.exists()]
            items.append((idx, v, match, room, context))
    return items


def _pdf_build_evidence_section(evidence_items: list[Any], styles, evidence_caption_style):
    from reportlab.lib.units import inch
    from reportlab.lib import colors as _colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, Image, KeepTogether, Table, TableStyle

    story: list[Any] = []
    for idx, v, img_path, room_path, context in evidence_items:
        ts_str = fmt_ist(v.get("created_at", ""))
        sev = v["severity"].upper()
        sev_color = "#c0392b" if v["severity"] == "high" else "#d68910"
        vtype_pretty = v["violation_type"].replace("_", " ").title()
        caption = (f'<b>#{idx} \u2014 {vtype_pretty}</b>  \xb7  '
                   f'<font color="{sev_color}"><b>{sev}</b></font>  \xb7  {ts_str}')
        detail = _pdf_clean_details(v.get("details"))
        if detail:
            caption += f'<br/><font size="8" color="#666">{detail}</font>'
        try:
            if room_path is not None:
                # Two cameras side by side: primary (webcam) + phone (room).
                cap = ParagraphStyle('evcap', parent=evidence_caption_style,
                                     fontSize=8, textColor=_colors.HexColor("#666"),
                                     alignment=1, spaceBefore=2)
                primary = Image(str(img_path), width=2.6 * inch, height=2.0 * inch, kind="proportional")
                phone = Image(str(room_path), width=2.6 * inch, height=2.0 * inch, kind="proportional")
                grid = Table(
                    [[primary, phone],
                     [Paragraph("Primary camera", cap), Paragraph("Phone camera", cap)]],
                    colWidths=[2.75 * inch, 2.75 * inch])
                grid.setStyle(TableStyle([
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]))
                story.append(KeepTogether([Paragraph(caption, evidence_caption_style), grid, Spacer(1, 14)]))
            else:
                img = Image(str(img_path), width=4.5 * inch, height=3.4 * inch, kind="proportional")
                story.append(KeepTogether([Paragraph(caption, evidence_caption_style), img, Spacer(1, 14)]))
            # Pre-violation context strip (oldest-first) — small thumbs showing
            # the seconds leading up to the flag, so context (dropped pen) is
            # distinguishable from intent (phone).
            if context:
                _ctxcap = ParagraphStyle('evctxcap', parent=evidence_caption_style,
                                         fontSize=7, textColor=_colors.HexColor("#888"),
                                         alignment=0, spaceBefore=2)
                _thumbs = [Image(str(p), width=1.3 * inch, height=1.0 * inch, kind="proportional")
                           for p in context]
                _ctx_tbl = Table([_thumbs], colWidths=[1.4 * inch] * len(_thumbs))
                _ctx_tbl.setStyle(TableStyle([
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 1),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ]))
                story.append(KeepTogether([
                    Paragraph("Context — seconds before the flag:", _ctxcap),
                    _ctx_tbl, Spacer(1, 14)]))
        except Exception as err:
            logger.warning("pdf: unreadable screenshot %s: %s", img_path, err)
            story.append(Paragraph(
                caption + f'  <font color="#999">(image unreadable: {err})</font>',
                evidence_caption_style))
            story.append(Spacer(1, 8))
    return story


@router.get("/api/v1/export-pdf/{session_id:path}")
@limiter.limit("10/minute")
async def export_pdf(session_id: str, request: Request):
    teacher = await require_admin(request)
    scope = await resolve_scope(teacher, request)
    sess = await assert_session_accessible(session_id, scope)  # 404s cross-tenant
    tid = str(sess["teacher_id"])  # the session OWNER's tid, not the caller's
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                         TableStyle, Paragraph, Spacer,
                                         Image, PageBreak, KeepTogether)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        exam = await _assert_session_owned(session_id, tid)
        exam["full_name"] = await resolve_student_name(exam, tid)
        raw_violations = await _pdf_fetch_violations(session_id, tid)
        answers = await _pdf_fetch_answers(session_id, tid)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        story: list[Any] = []

        story.append(Paragraph("AI Proctored Exam \u2014 Audit Report", styles["Title"]))
        story.append(Paragraph(
            "Full proctoring evidence log \u2014 per-event violations, confidence "
            "and captured screenshots. For review and appeal handling.",
            ParagraphStyle("auditsub", parent=styles["Normal"], fontSize=9,
                           textColor=colors.HexColor("#666666"), spaceAfter=6)))
        story.append(Spacer(1, 12))

        risk = await compute_risk_score(session_id, teacher_id=tid)
        story.append(_pdf_build_info_table(exam, raw_violations, risk))
        story.append(Spacer(1, 20))

        story.append(Paragraph(f"Violations ({len(raw_violations)} total)", styles["Heading2"]))
        story.append(Spacer(1, 8))

        if raw_violations:
            vt, conf_vals = _pdf_build_violations_table(raw_violations)
            story.append(vt)
            story.append(Spacer(1, 8))
            ct = _pdf_build_confidence_summary(conf_vals, len(raw_violations))
            if ct:
                story.append(ct)
        else:
            story.append(Paragraph("No violations recorded.", styles["Normal"]))

        evidence_items = _pdf_find_evidence(session_id, exam, raw_violations, tid)
        if evidence_items:
            story.append(Spacer(1, 18))
            story.append(PageBreak())
            story.append(Paragraph(
                f"Visual Evidence ({len(evidence_items)} captures)", styles["Heading2"]))
            story.append(Paragraph(
                "Screenshots are listed in the same order as the violations table above.",
                styles["Italic"]))
            story.append(Spacer(1, 10))

            evidence_caption_style = ParagraphStyle(
                "EvidenceCaption", parent=styles["Normal"],
                fontSize=9, leading=12, spaceAfter=4)
            story.extend(_pdf_build_evidence_section(evidence_items, styles, evidence_caption_style))

        story.append(Spacer(1, 20))
        story.append(Paragraph("Answer Sheet", styles["Heading2"]))
        story.append(Spacer(1, 8))

        try:
            pdf_questions = await _load_questions(teacher_id=tid, exam_id=exam.get("exam_id"))
            q_correct = {str(q["id"]): q["correct"] for q in pdf_questions}
            q_texts = {q["id"]: q.get("question", "")[:50] for q in pdf_questions}
        except Exception as e:
            logger.warning("Failed to load questions for PDF export: %s", e)
            q_correct = {}
            q_texts = {}

        if answers:
            ad = [["#", "Question", "Student", "Correct", "Result"]]
            for a in answers:
                qid = str(a["question_id"])
                correct = q_correct.get(qid, "?")
                is_right = str(a["answer"]) == correct
                ad.append([f"Q{qid}", q_texts.get(qid, "")[:40], a["answer"], correct, "\u2713" if is_right else "\u2717"])
            at = Table(ad, colWidths=[30, 180, 50, 50, 40])
            at.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8f9fa"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(at)

        story.append(Spacer(1, 20))
        story.append(Paragraph(
            f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')} IST | Session: {session_id[:20]}...",
            styles["Normal"]))

        doc.build(story)
        buf.seek(0)
        fname = (f"report_{_safe_filename(exam.get('roll_number') or 'unknown')}_"
                 f"{now_ist().strftime('%Y%m%d')}.pdf")
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}"})

    except HTTPException:
        raise
    except Exception as e:
        _admin_log.error("[PDF] %s", e)
        raise HTTPException(status_code=500, detail=f"PDF error: {e}")


@router.get("/api/v1/admin/scorecard-pdf/{session_id:path}")
@limiter.limit("10/minute")
async def scorecard_pdf(session_id: str, request: Request):
    teacher = await require_admin(request)
    scope = await resolve_scope(teacher, request)
    sess = await assert_session_accessible(session_id, scope)  # 404s cross-tenant
    tid = str(sess["teacher_id"])  # the session OWNER's tid, not the caller's
    try:
        pdf_bytes, fname, _ = await _build_scorecard_pdf(session_id, tid)
        return StreamingResponse(
            io.BytesIO(pdf_bytes), media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}"})
    except HTTPException:
        raise
    except Exception as e:
        _admin_log.error("[Scorecard PDF] %s", e)
        raise HTTPException(status_code=500, detail=f"Scorecard PDF error: {e}")


@router.get("/api/v1/admin/scorecard-zip")
@limiter.limit("5/minute")
async def scorecard_zip(request: Request, exam_id: Optional[str] = None):
    teacher = await require_admin(request)
    scope = await resolve_scope(teacher, request)
    tids = await scope_to_teacher_ids(scope)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                         TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet

        sess_q = _atable("exam_sessions")\
            .select("session_key,roll_number,full_name,score,total,percentage,time_taken_secs,risk_score,started_at,submitted_at,exam_id,teacher_id")\
            .in_("status", list(RESULT_STATUSES))
        if tids is not None:
            if not tids:
                sess_q = sess_q.eq("teacher_id", "__none__")  # empty org → match nothing
            elif len(tids) == 1:
                sess_q = sess_q.eq("teacher_id", str(tids[0]))
            else:
                sess_q = sess_q.in_("teacher_id", tids)
        # else (superadmin / None): no teacher filter
        if exam_id:
            sess_q = sess_q.eq("exam_id", exam_id)
        sessions = (await sess_q.execute()).data or []
        if not sessions:
            raise HTTPException(status_code=404, detail="No completed sessions found")

        eid = exam_id or (sessions[0].get("exam_id") if sessions else None)
        # Questions/config are exam-scoped, loaded per the owner of the
        # first session's exam. In a multi-teacher org the question bank can
        # differ per owner, so resolve them per-session below using each
        # row's own teacher_id.
        _q_cache: dict[tuple[str, str], list[dict[str, Any]]] = {}
        _cfg_cache: dict[tuple[str, str], dict[str, Any] | None] = {}

        async def _questions_for(owner_tid: str, ex_id):
            key = (owner_tid, ex_id)
            if key not in _q_cache:
                _q_cache[key] = await _load_questions(teacher_id=owner_tid, exam_id=ex_id)
            return _q_cache[key]

        async def _config_for(owner_tid: str, ex_id):
            key = (owner_tid, ex_id)
            if key not in _cfg_cache:
                cfg = None
                try:
                    cfg = await _load_exam_config(str(owner_tid), exam_id=ex_id)
                except Exception as e:
                    logger.debug("Failed to load exam config for ZIP export: %s", e)
                _cfg_cache[key] = cfg
            return _cfg_cache[key]
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for sess in sessions:
                sid = sess["session_key"]
                # Per-session OWNER tid — the loop can span sessions owned by
                # different teachers in an org, so answer/question loads MUST
                # use each row's own teacher_id, not the caller's.
                row_tid = str(sess.get("teacher_id"))
                row_eid = sess.get("exam_id") or eid
                questions = await _questions_for(row_tid, row_eid)
                config = await _config_for(row_tid, row_eid)
                exam_title = (config or {}).get("title", "Exam")
                ans_rows = (await _atable("answers").select("question_id,answer")
                            .eq("session_key", sid)
                            .eq("teacher_id", row_tid).execute()).data or []
                ans_map = {str(a["question_id"]): a["answer"] for a in ans_rows}

                pdf_buf = io.BytesIO()
                doc = SimpleDocTemplate(pdf_buf, pagesize=A4, topMargin=40, bottomMargin=40)
                styles = getSampleStyleSheet()
                story: list[Any] = []

                story.append(Paragraph(f"Scorecard — {exam_title}", styles["Title"]))
                story.append(Spacer(1, 12))

                score = sess.get("score", 0)
                total = sess.get("total", 0)
                pct = sess.get("percentage", 0)
                passed = pct >= ((config or {}).get("pass_mark") or 40)

                info = [
                    ["Field", "Value"],
                    ["Student Name", sess.get("full_name", "")],
                    ["Roll Number", sess.get("roll_number", "")],
                    ["Date", fmt_ist(sess.get("submitted_at", sess.get("started_at", "")))],
                    ["Score", f"{score}/{total}"],
                    ["Percentage", f"{pct}%"],
                    ["Result", "PASS" if passed else "FAIL"],
                    ["Time Taken", f"{sess.get('time_taken_secs', 0) // 60}m {sess.get('time_taken_secs', 0) % 60}s"],
                ]
                t = Table(info, colWidths=[140, 330])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
                    ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                    ("FONTNAME",  (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",  (0,0), (-1, -1), 10),
                    ("ROWBACKGROUNDS", (0,1), (-1, -1),
                     [colors.HexColor("#f0f4ff"), colors.white]),
                    ("GRID",    (0,0), (-1, -1), 0.5, colors.grey),
                    ("PADDING", (0,0), (-1, -1), 8),
                ]))
                story.append(t)
                story.append(Spacer(1, 20))

                story.append(Paragraph("Question-wise Results", styles["Heading2"]))
                story.append(Spacer(1, 8))

                if questions:
                    qd = [["#", "Question", "Your Answer", "Correct", "Result"]]
                    for i, q in enumerate(questions, 1):
                        qid = str(q.get("question_id", q.get("id", "")))
                        correct_ans = str(q.get("correct", ""))
                        student_ans = ans_map.get(qid, "\u2014")
                        is_right = str(student_ans) == correct_ans
                        q_text = q.get("question", "")
                        if len(q_text) > 60:
                            q_text = q_text[:57] + "..."
                        qd.append([str(i), q_text, str(student_ans)[:20], correct_ans[:20],
                                   "\u2713" if is_right else "\u2717"])
                    qt = Table(qd, colWidths=[25, 230, 80, 80, 35])
                    qt.setStyle(TableStyle([
                        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
                        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                        ("FONTNAME",  (0,0), (-1,0), "Helvetica-Bold"),
                        ("FONTSIZE",  (0,0), (-1, -1), 9),
                        ("ROWBACKGROUNDS", (0,1), (-1, -1),
                         [colors.HexColor("#f8f9fa"), colors.white]),
                        ("GRID",    (0,0), (-1, -1), 0.5, colors.grey),
                        ("PADDING", (0,0), (-1, -1), 6),
                        ("ALIGN", (4,1), (4, -1), "CENTER"),
                    ]))
                    story.append(qt)

                story.append(Spacer(1, 20))
                story.append(Paragraph(
                    f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')} IST",
                    styles["Normal"]))

                doc.build(story)
                pdf_buf.seek(0)
                roll = _safe_filename(sess.get("roll_number"), "unknown")
                zf.writestr(f"scorecard_{roll}.pdf", pdf_buf.getvalue())

        buf.seek(0)
        fname = f"scorecards_{_safe_filename(exam_id or 'all')}_{now_ist().strftime('%Y%m%d')}.zip"
        return StreamingResponse(
            buf, media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={fname}"})

    except HTTPException:
        raise
    except Exception as e:
        _admin_log.error("[Scorecard ZIP] %s", e)
        raise HTTPException(status_code=500, detail=f"Scorecard ZIP error: {e}")


@router.post("/api/v1/admin/exams/{exam_id}/email-scorecards")
@limiter.limit("5/minute")
async def email_scorecards(exam_id: str, request: Request, body: EmailScorecardsIn = Body(default=EmailScorecardsIn())):
    teacher = await require_admin(request)
    # Scope-aware (matches failed_sessions / export_csv): an org admin may email
    # scorecards for a co-teacher's exam. Reads roll up org-wide; the per-session
    # claim + enqueue key on each session's OWNER teacher_id (a single exam_id is
    # owned by one teacher) so the job is filed under the owner and the student
    # sees their own teacher's name — not the acting admin's.
    scope = await resolve_scope(teacher, request)
    tids = await scope_to_teacher_ids(scope)

    resend_all = body.resend_all
    custom_message = body.custom_message.strip() or None
    actor_name = teacher.get("full_name") or teacher.get("email") or "Your teacher"

    # Build the query (NOT awaited — the PostgresTable chain is not awaitable;
    # only .execute() is). The stray `await` here was the email-scorecards 500:
    # "object PostgresTable can't be used in 'await' expression".
    sess_q = (apply_teacher_scope(_atable("exam_sessions").select(
        "session_key,roll_number,full_name,exam_id,scorecard_emailed_at,teacher_id"
    ), tids).in_("status", list(RESULT_STATUSES)).eq("exam_id", exam_id)
        .limit(1000))
    sessions = (await sess_q.execute()).data or []
    if not sessions:
        raise HTTPException(status_code=404, detail="No completed sessions found for this exam")

    # Resolve each owner teacher's display name once so the email signs off as
    # the student's real teacher even when an admin triggers the send.
    owner_names: dict[str, str] = {}
    owner_tids = sorted({str(s.get("teacher_id") or "") for s in sessions if s.get("teacher_id")})
    if owner_tids:
        try:
            t_rows = (await _atable("teachers").select("id,full_name,email")
                      .in_("id", owner_tids).execute()).data or []
            for t in t_rows:
                owner_names[str(t["id"])] = t.get("full_name") or t.get("email") or actor_name
        except Exception as e:
            _admin_log.warning("[email-scorecards] owner-name lookup failed: %s", e)

    roll_emails: dict[str, str] = {}
    try:
        inv_rows = (await apply_teacher_scope(_atable("student_invites").select("roll_number,email"), tids)
                    .eq("exam_id", exam_id).execute()).data or []
        for r in inv_rows:
            roll = str(r.get("roll_number") or "").strip().upper()
            email = str(r.get("email") or "").strip().lower()
            if roll and email:
                roll_emails[roll] = email
    except Exception as e:
        _admin_log.warning("[email-scorecards] invite lookup failed: %s", e)
    try:
        stud_rows = (await apply_teacher_scope(_atable("students").select("roll_number,email"), tids)
                     .execute()).data or []
        for r in stud_rows:
            roll = str(r.get("roll_number") or "").strip().upper()
            email = str(r.get("email") or "").strip().lower()
            if roll and email and roll not in roll_emails:
                roll_emails[roll] = email
    except Exception as e:
        _admin_log.warning("[email-scorecards] student lookup failed: %s", e)

    sent = 0
    failed = 0
    already_sent = 0
    skipped_no_email = 0
    failures: list[dict[str, Any]] = []

    async def _release_claim(sid: str, owner_tid: str) -> None:
        """Re-open a claimed-but-unsent scorecard so it can be retried without a
        blanket resend_all. The claim sets scorecard_emailed_at=now BEFORE the
        email is enqueued/sent; if that step fails, leaving the timestamp set
        would permanently mark the scorecard 'sent' (future non-resend runs skip
        it), silently losing it. Best-effort rollback to NULL."""
        try:
            await _atable("exam_sessions").update({"scorecard_emailed_at": None})\
                .eq("session_key", sid).eq("teacher_id", owner_tid).execute()
        except Exception:
            _admin_log.warning("[email-scorecards] claim rollback failed for %s", sid)

    for sess in sessions:
        sid = sess["session_key"]
        roll = str(sess.get("roll_number") or "").strip().upper()
        full_name = sess.get("full_name") or "Student"
        owner_tid = str(sess.get("teacher_id") or "")
        teacher_name = owner_names.get(owner_tid, actor_name)

        if sess.get("scorecard_emailed_at") and not resend_all:
            already_sent += 1
            continue

        student_email_val: str | None = roll_emails.get(roll)
        if not student_email_val:
            skipped_no_email += 1
            failures.append({"roll": roll, "reason": "no email on file"})
            continue

        if not resend_all:
            claim = await _atable("exam_sessions").update({
                "scorecard_emailed_at": now_ist().isoformat(),
            }).eq("session_key", sid).eq("teacher_id", owner_tid).is_("scorecard_emailed_at", "null").execute()
            claimed = bool(claim.data)
            if not claimed:
                already_sent += 1
                continue

        try:
            job_result = enqueue_job(
                send_scorecard_email_job,
                session_key=sid,
                teacher_id=owner_tid,
                email=email,
                full_name=full_name,
                teacher_name=teacher_name,
                custom_message=custom_message,
                resend_all=resend_all,
            )
        except Exception as e:
            # One bad session must not 500 the whole batch (the endpoint had no
            # error handling — a single enqueue/job failure took the lot down).
            # Record it and log the traceback so it's diagnosable next time.
            # Roll the claim back so this scorecard can be retried without a
            # blanket resend_all (the claim already marked it emailed=now).
            await _release_claim(sid, owner_tid)
            failed += 1
            failures.append({"roll": roll, "reason": f"enqueue failed: {e}"})
            _admin_log.error("[email-scorecards] enqueue failed for %s: %s",
                             sid, e, exc_info=True)
            continue
        if job_result is None:
            sent += 1
        elif job_result.get("ok"):
            sent += 1
        else:
            # Inline send (RQ disabled) reported failure — re-open the claim so
            # a retry isn't blocked by the now-set emailed_at timestamp.
            await _release_claim(sid, owner_tid)
            failed += 1
            failures.append({"roll": roll, "reason": job_result.get("error", "send failed")})

    return {
        "sent": sent,
        "failed": failed,
        "already_sent": already_sent,
        "skipped_no_email": skipped_no_email,
        "total": len(sessions),
        "failures": failures[:50],
    }


@router.get("/api/v1/admin-failed-sessions")
@limiter.limit("30/minute")
async def failed_sessions(request: Request, exam_id: Optional[str] = None):
    # Org-rollup, scope-aware (matches export_csv / scorecard_zip): an org admin
    # sees co-teachers' failed submissions; a plain teacher stays own-scoped.
    teacher = await require_admin(request)
    scope = await resolve_scope(teacher, request)
    tids = await scope_to_teacher_ids(scope)
    failed = await apply_teacher_scope(
        _atable("violations").select("session_key").eq("violation_type", "submit_failed"),
        tids).execute()
    failed_keys = {r["session_key"] for r in (failed.data or [])}
    sub_query = apply_teacher_scope(
        _atable("exam_sessions").select("session_key")
            .in_("status", list(RESULT_STATUSES))
            .in_("session_key", list(failed_keys) or ["__none__"]),
        tids)
    if exam_id:
        sub_query = sub_query.eq("exam_id", exam_id)
    submitted = await sub_query.execute()
    submitted_keys = {r["session_key"] for r in (submitted.data or [])}
    if exam_id:
        es = await apply_teacher_scope(
            _atable("exam_sessions").select("session_key").eq("exam_id", exam_id),
            tids).execute()
        exam_skeys = {r["session_key"] for r in (es.data or [])}
        failed_keys = failed_keys & exam_skeys
    unrecovered = [k for k in failed_keys if k not in submitted_keys]
    return {"failed_sessions": unrecovered, "count": len(unrecovered)}


__all__ = ["router"]
