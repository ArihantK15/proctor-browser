"""Admin router — all teacher-facing endpoints.

Extracted from main.py. Imports shared dependencies from `dependencies`.
"""

import io
import csv
import json
import base64
import hashlib
import time
import zipfile
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Request, HTTPException, Body
from pydantic import BaseModel, ConfigDict

from ..dependencies import (
    supabase, get_logger, fmt_ist, require_admin, require_auth,
    now_ist, _assert_session_owned, _load_exam_config, _load_questions,
    _recalculate_score, _safe_filename,
    compute_risk_score, _build_sessions_payload, _partition_live_sessions,
    _clear_token_issue, _clear_token_consume, _CLEAR_TOKEN_TTL, _CLEAR_ACTIVE_WINDOW,
    SCREENSHOTS_DIR, _cache, _atable,
    _collect_session_screenshots, _is_violation, _match_screenshot_for_violation,
    _get_invite_base_url, _get_teacher_by_id,
    INVITE_DAILY_CAP, _new_invite_token, _uuid,
    _safe_path_component, _assert_within_directory, _html_escape,
    SessionStatus, InviteStatus, VerificationStatus,
    SECRET_KEY,
)

router = APIRouter(prefix="")


# ─── PYDANTIC MODELS ─────────────────────────────────

class IdDecisionIn(BaseModel):
    model_config = ConfigDict(strict=True)
    violation_id: int
    session_key: str
    decision: str  # "approved" | "retake" | "rejected"


# ─── HELPER: BUILD SCORECARD PDF ─────────────────────────

def _build_scorecard_pdf(session_id: str, teacher_id) -> tuple[bytes, str, dict]:
    """Render a single student's scorecard as a PDF and return
    ``(bytes, filename, exam_summary)``.

    Centralised so the /scorecard-pdf endpoint, the bulk ZIP route,
    and the "email scorecards to students" flow all produce the exact
    same document. ``teacher_id`` must already be validated — this
    helper is trusted-internal; it assumes ownership was asserted by
    the caller.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table,
                                     TableStyle, Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet

    tid = teacher_id
    exam = _assert_session_owned(session_id, tid)
    exam_id = exam.get("exam_id")

    questions = _load_questions(teacher_id=tid, exam_id=exam_id)
    ans_rows = (supabase.table("answers").select("question_id,answer")
                .eq("session_key", session_id)
                .eq("teacher_id", str(tid)).execute()).data or []
    ans_map = {str(a["question_id"]): a["answer"] for a in ans_rows}

    config = None
    try:
        config = _load_exam_config(str(tid), exam_id=exam_id)
    except Exception:
        pass
    exam_title = (config or {}).get("exam_title") or (config or {}).get("title") or "Exam"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"Scorecard — {exam_title}", styles["Title"]))
    story.append(Spacer(1, 12))

    score = exam.get("score", 0)
    total = exam.get("total", 0)
    pct = exam.get("percentage", 0)
    risk = compute_risk_score(session_id, teacher_id=tid)
    passed = pct >= 40

    info = [
        ["Field", "Value"],
        ["Student Name", exam.get("full_name", "")],
        ["Roll Number", exam.get("roll_number", "")],
        ["Date", fmt_ist(exam.get("submitted_at", exam.get("started_at", "")))],
        ["Score", f"{score}/{total}"],
        ["Percentage", f"{pct}%"],
        ["Result", "PASS" if passed else "FAIL"],
        ["Time Taken", f"{exam.get('time_taken_secs', 0) // 60}m {exam.get('time_taken_secs', 0) % 60}s"],
        ["Risk Level", risk["label"]],
    ]
    t = Table(info, colWidths=[140, 330])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.HexColor("#f0f4ff"), colors.white]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("PADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))

    # ── Violation Summary ────────────────────────────────────────────
    viol_rows = (supabase.table("violations")
                 .select("violation_type, severity")
                 .eq("session_key", session_id)
                 .eq("teacher_id", str(tid)).execute()).data or []

    viol_counts: dict[str, dict[str, int]] = {}
    for v in viol_rows:
        vtype = v.get("violation_type", "unknown")
        sev = v.get("severity", "low")
        if vtype not in viol_counts:
            viol_counts[vtype] = {"high": 0, "medium": 0, "low": 0, "total": 0}
        viol_counts[vtype][sev] = viol_counts[vtype].get(sev, 0) + 1
        viol_counts[vtype]["total"] += 1

    HUMAN_NAMES: dict[str, str] = {
        "gaze_away": "Gaze Away",
        "head_turned": "Head Turned",
        "eyes_closed": "Eyes Closed",
        "face_missing": "Face Missing",
        "multiple_faces": "Multiple Faces",
        "wrong_person": "Wrong Person",
        "calibration_abort": "Calibration Aborted (Identity Swap)",
        "cheat_object_detected": "Cheat Object Detected",
        "voice_detected": "Voice Detected",
        "window_focus_lost": "Window Focus Lost",
        "tab_hidden": "Tab Hidden",
        "shortcut_blocked": "Shortcut Blocked",
        "vm_detected": "VM Detected",
        "remote_desktop_detected": "Remote Desktop",
        "screen_share_detected": "Screen Share",
        "multiple_monitors": "Multiple Monitors",
        "phone_consulting": "Phone Consulting (Behavioral)",
        "collaboration": "Collaboration Suspected (Behavioral)",
        "answer_memo": "Answer Memorization (Behavioral)",
        "note_reading": "Note Reading (Behavioral)",
        "sustained_offtask": "Sustained Off-Task (Behavioral)",
        "nervous_evasion": "Nervous Evasion (Behavioral)",
    }

    if viol_counts:
        story.append(Paragraph("Violation Summary", styles["Heading2"]))
        story.append(Spacer(1, 8))

        vdata = [["Violation", "Count"]]
        for vtype in sorted(viol_counts.keys(), key=lambda k: viol_counts[k]["total"], reverse=True):
            name = HUMAN_NAMES.get(vtype, vtype.replace("_", " ").title())
            vdata.append([name, str(viol_counts[vtype]["total"])])

        vtotal = ["Total", str(sum(v["total"] for v in viol_counts.values()))]
        vdata.append(vtotal)

        vt = Table(vdata, colWidths=[370, 100])
        row_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("PADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ]
        for i in range(1, len(vdata)):
            bg = colors.HexColor("#f0f4ff") if i % 2 == 1 else colors.white
            row_style.append(("ROWBACKGROUNDS", (0, i), (-1, i), bg))
        row_style.append(("BACKGROUND", (0, len(vdata) - 1), (-1, len(vdata) - 1),
                          colors.HexColor("#e8ecf4")))
        row_style.append(("FONTNAME", (0, len(vdata) - 1), (-1, len(vdata) - 1),
                          "Helvetica-Bold"))
        vt.setStyle(TableStyle(row_style))
        story.append(vt)
        story.append(Spacer(1, 20))

    story.append(Paragraph("Question-wise Results", styles["Heading2"]))
    story.append(Spacer(1, 8))

    if questions:
        qd = [["#", "Question", "Your Answer", "Correct Answer", "Result"]]
        for i, q in enumerate(questions, 1):
            qid = str(q.get("question_id", q.get("id", "")))
            correct_ans = str(q.get("correct", ""))
            student_ans = ans_map.get(qid, "—")
            is_right = str(student_ans) == correct_ans
            q_text = q.get("question", "")
            if len(q_text) > 60:
                q_text = q_text[:57] + "..."
            qd.append([
                str(i),
                q_text,
                str(student_ans)[:20],
                correct_ans[:20],
                "✓" if is_right else "✗",
            ])
        qt = Table(qd, colWidths=[25, 230, 80, 80, 35])
        qt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.HexColor("#f8f9fa"), colors.white]),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("PADDING", (0,0), (-1,-1), 6),
            ("ALIGN", (4,1), (4,-1), "CENTER"),
        ]))
        story.append(qt)
    else:
        story.append(Paragraph("No questions available.", styles["Normal"]))

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')} IST",
        styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    roll = _safe_filename(exam.get("roll_number"), "unknown")
    fname = f"scorecard_{roll}_{now_ist().strftime('%Y%m%d')}.pdf"
    summary = {
        "exam": exam,
        "exam_title": exam_title,
        "score": score,
        "total": total,
        "percentage": pct,
        "passed": passed,
        "risk_label": risk["label"],
        "violations": viol_counts,
        "total_violations": sum(v["total"] for v in viol_counts.values()),
    }
    return buf.getvalue(), fname, summary

# ═══════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════

# ─── 1. PENDING ID VERIFICATIONS ─────────────────────────

@router.get("/api/v1/admin/pending-verifications")
def pending_verifications(request: Request, exam_id: str = None):
    """Return all pending ID verifications for this teacher."""
    teacher = require_admin(request)
    tid = teacher["id"]
    import json as _json
    query = supabase.table("violations")\
        .select("*")\
        .eq("teacher_id", str(tid))\
        .eq("violation_type", "id_verification")\
        .order("created_at", desc=True)
    result = query.execute()

    legacy_session_keys = None
    if exam_id:
        es = supabase.table("exam_sessions").select("session_key")\
            .eq("teacher_id", str(tid)).eq("exam_id", exam_id).execute()
        legacy_session_keys = {r["session_key"] for r in (es.data or [])}

    pending = []
    for row in (result.data or []):
        try:
            obj = json.loads(row.get("details", "{}"))
        except Exception:
            continue
        if obj.get("status") != VerificationStatus.PENDING:
            continue
        if exam_id:
            stamped_eid = obj.get("exam_id") or ""
            if stamped_eid:
                if stamped_eid != exam_id:
                    continue
            else:
                if row.get("session_key") not in (legacy_session_keys or set()):
                    continue
        roll = obj.get("roll_number", "")
        pending.append({
            "id":           row.get("id"),
            "session_key":  row.get("session_key"),
            "roll_number":  roll,
            "full_name":    obj.get("full_name", ""),
            "selfie_url":   f"/api/v1/admin/screenshot/{roll}/{obj['selfie_file']}"
                            if obj.get("selfie_file") else None,
            "id_url":       f"/api/v1/admin/screenshot/{roll}/{obj['id_file']}"
                            if obj.get("id_file") else None,
            "created_at":   fmt_ist(row.get("created_at", "")),
        })
    return {"pending": pending}


# ─── 2. ID DECISION ─────────────────────────────────

@router.post("/api/v1/admin/id-decision")
def id_decision(data: IdDecisionIn, request: Request):
    """Teacher approves, requests retake, or rejects a student's ID."""
    teacher = require_admin(request)
    tid = teacher["id"]
    if data.decision not in ("approved", "retake", "rejected"):
        raise HTTPException(status_code=400, detail="Invalid decision")
    import json as _json
    result = supabase.table("violations")\
        .select("*")\
        .eq("id", data.violation_id)\
        .eq("teacher_id", str(tid))\
        .limit(1)\
        .execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Verification not found")
    row = result.data[0]
    try:
        obj = json.loads(row.get("details", "{}"))
    except Exception:
        obj = {}
    obj["status"] = data.decision
    obj["decided_by"] = teacher.get("full_name", teacher.get("email", ""))
    obj["decided_at"] = now_ist().isoformat()
    supabase.table("violations")\
        .update({"details": json.dumps(obj)})\
        .eq("id", data.violation_id)\
        .execute()

    if data.decision == "rejected":
        reject_row = {
            "session_key":    data.session_key,
            "violation_type": "id_rejected",
            "severity":       "high",
            "details":        f"Teacher rejected student identity — "
                              f"decided by {obj['decided_by']}",
        }
        if tid:
            reject_row["teacher_id"] = str(tid)
        supabase.table("violations").insert(reject_row).execute()
        if _cache:
            _cache.delete(f"risk_score:{data.session_key}")
        try:
            supabase.table("exam_sessions").update({
                "status":       SessionStatus.REJECTED,
                "submitted_at": now_ist().isoformat(),
            }).eq("session_key", data.session_key).execute()
        except Exception:
            pass

    return {"status": "ok", "decision": data.decision}


# ─── 3. RISK SCORE ─────────────────────────────────

@router.get("/api/v1/risk-score/{session_id:path}")
def get_risk_score(session_id: str, request: Request):
    """Compute behavioral risk score for any session (live or completed)."""
    teacher = require_admin(request)
    tid = teacher["id"]
    _assert_session_owned(session_id, tid)
    result = compute_risk_score(session_id, teacher_id=tid)
    result["session_id"] = session_id
    return result

# ─── 4. TIMELINE ─────────────────────────────────

@router.get("/api/v1/admin/timeline/{session_id:path}")
def get_timeline(session_id: str, request: Request):
    """Full forensics timeline: every event + screenshot paths for a session."""
    teacher = require_admin(request)
    tid = teacher["id"]
    session_info = _assert_session_owned(session_id, tid)
    viol_result = supabase.table("violations")\
        .select("*")\
        .eq("session_key", session_id)\
        .eq("teacher_id", str(tid))\
        .order("created_at")\
        .execute()
    events = viol_result.data or []

    roll = session_info.get("roll_number") or (
        session_id.rsplit("_", 1)[0] if "_" in session_id else session_id[:20]
    )
    screenshot_paths = _collect_session_screenshots(roll, str(tid))
    screenshot_urls = {
        fname: f"/api/v1/admin/screenshot/{roll}/{fname}"
        for fname in screenshot_paths
    }

    timeline = []
    for e in events:
        entry = {
            "id":        e.get("id"),
            "type":      e["violation_type"],
            "severity":  e["severity"],
            "timestamp": fmt_ist(e.get("created_at", "")),
            "raw_ts":    e.get("created_at", ""),
            "details":   e.get("details"),
            "is_violation": _is_violation(e["violation_type"]),
        }
        match = _match_screenshot_for_violation(e, screenshot_paths)
        if match is not None:
            entry["screenshot"] = screenshot_urls[match.name]
        timeline.append(entry)

    return {
        "session_id":  session_id,
        "roll_number": session_info.get("roll_number", roll),
        "full_name":   session_info.get("full_name", ""),
        "status":      session_info.get("status", "unknown"),
        "started_at":  fmt_ist(session_info.get("started_at", "")),
        "submitted_at": fmt_ist(session_info.get("submitted_at", "")),
        "score":       session_info.get("score"),
        "total":       session_info.get("total"),
        "risk_score":  session_info.get("risk_score"),
        "total_events": len(events),
        "timeline":    timeline,
        "screenshots": list(screenshot_urls.values()),
    }


# ─── 5. UPLOAD QUESTION IMAGE ─────────────────────────

@router.post("/api/v1/admin/upload-question-image")
def upload_question_image(request: Request, body: dict = Body(...)):
    """Teacher uploads a question image."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    raw = body.get("image") or body.get("data") or ""
    if not isinstance(raw, str) or not raw:
        raise HTTPException(status_code=400, detail="Missing 'image' (base64)")
    if raw.startswith("data:"):
        try:
            _, raw = raw.split(",", 1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Malformed data URL")
    try:
        blob = base64.b64decode(raw, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 image payload")
    if len(blob) > 4 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Image too large (max 4MB)")

    if blob[:8] == b"\x89PNG\r\n\x1a\n":
        ext = "png"
        media = "image/png"
    elif blob[:3] == b"\xff\xd8\xff":
        ext = "jpg"
        media = "image/jpeg"
    elif blob[:6] in (b"GIF87a", b"GIF89a"):
        ext = "gif"
        media = "image/gif"
    elif blob[:4] == b"RIFF" and blob[8:12] == b"WEBP":
        ext = "webp"
        media = "image/webp"
    else:
        raise HTTPException(status_code=400, detail="Unsupported image format (PNG/JPEG/GIF/WebP only)")

    digest = hashlib.sha1(blob).hexdigest()[:24]
    filename = f"{digest}.{ext}"
    tdir = Path(QUESTION_IMG_DIR) / tid
    tdir.mkdir(parents=True, exist_ok=True)
    fpath = tdir / filename
    if not fpath.exists():
        try:
            with open(fpath, "wb") as f:
                f.write(blob)
        except OSError as e:
            print(f"[QImage] write failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to store image")

    url = f"/api/v1/question-image/{tid}/{filename}"
    return {"url": url, "bytes": len(blob), "media_type": media}
# ─── 6. SERVE QUESTION IMAGE ────────────────────────────

@router.get("/api/v1/question-image/{tid}/{filename}")
def get_question_image(tid: str, filename: str, request: Request):
    """Serve a question image."""
    from jose import jwt, JWTError
    auth = request.headers.get("Authorization", "")
    allowed = False
    if auth.startswith("Bearer "):
        tok = auth[7:]
        try:
            teacher = verify_admin_token(tok)
            if str(teacher.get("id")) == str(tid):
                allowed = True
        except HTTPException:
            pass
        if not allowed:
            try:
                payload = jwt.decode(
                    tok, SECRET_KEY, algorithms=["HS256"],
                    options={"verify_aud": False, "require": ["exp"]},
                )
                if str(payload.get("tid") or "") == str(tid):
                    allowed = True
            except JWTError:
                pass
    if not allowed:
        raise HTTPException(status_code=401, detail="Authentication required")

    safe_tid = _safe_path_component(tid)
    safe_file = _safe_path_component(filename)
    fpath = Path(QUESTION_IMG_DIR) / safe_tid / safe_file
    try:
        _assert_within_directory(fpath, Path(QUESTION_IMG_DIR))
    except (ValueError, RuntimeError):
        raise HTTPException(status_code=404, detail="Image not found")
    if not fpath.exists() or not fpath.is_file():
        raise HTTPException(status_code=404, detail="Image not found")
    suffix = fpath.suffix.lower()
    media_map = {".png": "image/png", ".jpg": "image/jpeg",
                 ".jpeg": "image/jpeg", ".gif": "image/gif",
                 ".webp": "image/webp"}
    media = media_map.get(suffix, "application/octet-stream")
    return FileResponse(str(fpath), media_type=media)


# ─── 7. SERVE SCREENSHOT ──────────────────────────────

@router.get("/api/v1/admin/screenshot/{roll}/{filename}")
def get_screenshot(roll: str, filename: str, request: Request):
    """Serve a screenshot image to the admin dashboard."""
    teacher = require_admin(request)
    safe_roll = _safe_path_component(roll)
    safe_file = _safe_path_component(filename)
    tid = str(teacher["id"])
    fpath = Path(SCREENSHOTS_DIR) / tid / safe_roll / safe_file
    try:
        _assert_within_directory(fpath, Path(SCREENSHOTS_DIR) / tid)
    except (ValueError, RuntimeError):
        raise HTTPException(status_code=404, detail="Screenshot not found")
    if not fpath.exists() or not fpath.is_file():
        raise HTTPException(status_code=404, detail="Screenshot not found")
    suffix = fpath.suffix.lower()
    media = "image/jpeg" if suffix in (".jpg", ".jpeg") else "image/png"
    return FileResponse(str(fpath), media_type=media,
                        headers={"Cache-Control": "private, max-age=3600"})


# ─── 8. LIVE SESSIONS VIEW ──────────────────────────────

@router.get("/api/v1/admin/sessions")
def get_all_sessions(request: Request, exam_id: str = None):
    """REST view of the Live tab."""
    teacher = require_admin(request)
    tid = teacher["id"]
    try:
        return _build_sessions_payload(str(tid), exam_id=exam_id)
    except Exception as e:
        print(f"[Sessions] ERROR: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ─── 9. RESULTS ─────────────────────────────────────────

@router.get("/api/v1/results")
def get_all_results(request: Request, exam_id: str = None):
    teacher = require_admin(request)
    return {"results": _fetch_all_results(teacher["id"], exam_id=exam_id)}


# ─── 10. EXPORT CSV ──────────────────────────────────────

@router.get("/api/v1/export-csv")
def export_csv(request: Request, exam_id: str = None):
    teacher = require_admin(request)
    results = _fetch_all_results(teacher["id"], exam_id=exam_id)
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["Timestamp","SessionID","RollNumber","FullName","Email",
                "Score","Total","Percentage","TimeTaken","Violations","RiskScore","RiskLabel"])
    for s in results:
        w.writerow([
            s["submitted_at"],
            s["session_id"],
            s["roll_number"],
            s["full_name"],
            s["email"],
            s["score"],
            s["total"],
            f"{s['percentage']}%",
            f"{s['time_taken_secs']}s",
            s["violation_count"],
            s.get("risk_score", ""),
            s.get("risk_label", ""),
        ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=results.csv"})


# ─── 11. EXPORT EXCEL ─────────────────────────────────────

@router.get("/api/v1/export-excel")
def export_excel(request: Request, exam_id: str = None):
    """Results export as a formatted .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    teacher = require_admin(request)
    results = _fetch_all_results(teacher["id"], exam_id=exam_id)

    wb = Workbook()
    ws = wb.active
    safe_eid = "".join(c for c in (exam_id or "all") if c.isalnum() or c in "-_")[:24]
    ws.title = f"Results_{safe_eid}" if safe_eid else "Results"

    headers = ["Timestamp", "Session ID", "Roll Number", "Full Name",
               "Email", "Score", "Total", "Percentage", "Time (min)",
               "Violations", "Risk Score", "Risk Label"]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    risk_fills = {
        "Low":    PatternFill("solid", fgColor="D1FAE5"),
        "Medium": PatternFill("solid", fgColor="FEF3C7"),
        "High":   PatternFill("solid", fgColor="FEE2E2"),
    }

    for s in results:
        try:
            pct_val = float(s.get("percentage") or 0)
        except Exception:
            pct_val = 0.0
        try:
            secs = int(s.get("time_taken_secs") or 0)
        except Exception:
            secs = 0
        mins = round(secs / 60, 2) if secs else 0

        ws.append([
            _xlsx_safe(s.get("submitted_at", "")),
            _xlsx_safe(s.get("session_id", "")),
            _xlsx_safe(s.get("roll_number", "")),
            _xlsx_safe(s.get("full_name", "")),
            _xlsx_safe(s.get("email", "")),
            s.get("score", 0),
            s.get("total", 0),
            pct_val,
            mins,
            s.get("violation_count", 0),
            s.get("risk_score", ""),
            _xlsx_safe(s.get("risk_label", "")),
        ])
        label = s.get("risk_label")
        fill = risk_fills.get(label)
        if fill:
            ws.cell(row=ws.max_row, column=12).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row,1)}"

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=8).number_format = '0.0"%"'
        ws.cell(row=row, column=9).number_format = '0.00'

    widths = [0] * len(headers)
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row):
            widths[i] = max(widths[i], min(len(str(v or "")), 40))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(w + 2, 10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"results_{safe_eid or 'all'}_{now_ist().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})
# ─── 12. EXPORT PDF ──────────────────────────────────

@router.get("/api/v1/export-pdf/{session_id:path}")
def export_pdf(session_id: str, request: Request):
    teacher = require_admin(request)
    tid = teacher["id"]
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                         TableStyle, Paragraph, Spacer,
                                         Image, PageBreak, KeepTogether)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        exam = _assert_session_owned(session_id, tid)

        viol_result = supabase.table("violations")\
            .select("*")\
            .eq("session_key", session_id)\
            .eq("teacher_id", str(tid))\
            .order("created_at").execute()
        raw_violations = [
            v for v in (viol_result.data or [])
            if v["severity"] in ("high", "medium")
        ]

        ans_result = supabase.table("answers")\
            .select("*")\
            .eq("session_key", session_id)\
            .eq("teacher_id", str(tid))\
            .execute()
        answers = ans_result.data or []

        buf    = io.BytesIO()
        doc    = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph("AI Proctored Exam — Report", styles["Title"]))
        story.append(Spacer(1, 12))

        info = [
            ["Field",          "Value"],
            ["Full Name",      exam["full_name"]],
            ["Roll Number",    exam["roll_number"]],
            ["Email",          exam.get("email", "")],
            ["Submitted At",   fmt_ist(exam.get("submitted_at", ""))],
            ["Score",          f"{exam.get('score',0)}/{exam.get('total',0)} "
                               f"({exam.get('percentage',0)}%)"],
            ["Time Taken",     f"{exam.get('time_taken_secs',0)} seconds "
                               f"({exam.get('time_taken_secs',0)//60}m "
                               f"{exam.get('time_taken_secs',0)%60}s)"],
            ["Total Violations", str(len(raw_violations))],
        ]
        risk = compute_risk_score(session_id, teacher_id=tid)
        info.append(["Behavioral Risk Score",
                     f"{risk['risk_score']}/100 — {risk['label']}"])
        t = Table(info, colWidths=[160, 310])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 10),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.HexColor("#f0f4ff"), colors.white]),
            ("GRID",    (0,0), (-1,-1), 0.5, colors.grey),
            ("PADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(t)
        story.append(Spacer(1, 20))

        story.append(Paragraph(
            f"Violations ({len(raw_violations)} total)", styles["Heading2"]))
        story.append(Spacer(1, 8))

        CONF_MAP = {
            "face_missing": 0.95, "multiple_faces": 0.92,
            "wrong_person": 0.78, "eyes_closed": 0.88,
            "earphone_detected": 0.72, "cheat_object_detected": 0.85,
            "gaze_away": 0.70, "head_away": 0.80,
            "voice_detected": 0.75, "window_focus_lost": 0.99,
            "tab_hidden": 0.99, "lighting_issue": 0.90,
            "shortcut_blocked": 0.99, "camera_covered": 0.95,
        }

        def get_conf(vtype, details):
            det = str(details or "")
            if "confidence:" in det:
                try:
                    raw = det.split("confidence:")[1].split("|")[0].strip()
                    return raw if "%" in raw else f"{raw}%"
                except Exception:
                    pass
            if "conf:" in det:
                try:
                    raw = det.split("conf:")[1].split(" ")[0].strip()
                    val = float(raw)
                    return f"{int(val)}%" if val > 1 else f"{int(val*100)}%"
                except Exception:
                    pass
            return f"{int(CONF_MAP.get(vtype, 0.75) * 100)}%"

        def clean_details(details):
            det = str(details or "")
            raw = det.split("| confidence:")[0].strip()[:40] \
                   if "| confidence:" in det else det[:40]
            return _html_escape(raw)

        if raw_violations:
            total_conf_vals = []
            vd = [["#", "Type", "Severity", "Time", "Conf", "Details"]]
            for i, v in enumerate(raw_violations, 1):
                conf_str = get_conf(v["violation_type"], v.get("details"))
                try:
                    total_conf_vals.append(float(conf_str.strip("%")) / 100)
                except Exception:
                    pass
                ts_part = ""
                if v.get("created_at"):
                    _fmted = fmt_ist(v["created_at"])
                    _comma = _fmted.find(",")
                    if _comma >= 0:
                        ts_part = _fmted[_comma+1:].replace("IST","").strip()
                    else:
                        ts_part = _fmted
                vd.append([
                    str(i),
                    v["violation_type"].replace("_", " ").title()[:22],
                    v["severity"].upper(),
                    ts_part,
                    conf_str,
                    clean_details(v.get("details"))[:35],
                ])
            vt = Table(vd, colWidths=[20, 120, 55, 70, 40, 165])
            vt.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#c0392b")),
                ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS", (0,1), (-1,-1),
                 [colors.HexColor("#fff5f5"), colors.white]),
                ("GRID",    (0,0), (-1,-1), 0.5, colors.grey),
                ("PADDING", (0,0), (-1,-1), 5),
                ("ALIGN",   (0,0), (0,-1), "CENTER"),
            ]))
            story.append(vt)
            story.append(Spacer(1, 8))

            if total_conf_vals:
                avg_conf  = sum(total_conf_vals) / len(total_conf_vals)
                high_conf = len([c for c in total_conf_vals if c >= 0.85])
                conf_data = [[
                    f"Overall Detection Confidence: {avg_conf:.0%}",
                    f"High Confidence Violations: {high_conf}/{len(raw_violations)}",
                    f"Reliability: {'High' if avg_conf>=0.85 else 'Medium' if avg_conf>=0.70 else 'Low'}",
                ]]
                ct = Table(conf_data, colWidths=[160, 160, 150])
                ct.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#eafaf1")),
                    ("TEXTCOLOR",  (0,0), (-1,-1), colors.HexColor("#1e8449")),
                    ("FONTNAME",   (0,0), (-1,-1), "Helvetica-Bold"),
                    ("FONTSIZE",   (0,0), (-1,-1), 8),
                    ("GRID",    (0,0), (-1,-1), 0.5, colors.grey),
                    ("PADDING", (0,0), (-1,-1), 6),
                    ("ALIGN",   (0,0), (-1,-1), "CENTER"),
                ]))
                story.append(ct)
        else:
            story.append(Paragraph("No violations recorded.", styles["Normal"]))

        # Visual Evidence Timeline
        roll_for_evidence = exam.get("roll_number") or (
            session_id.rsplit("_", 1)[0] if "_" in session_id else session_id[:20]
        )
        evidence_paths = _collect_session_screenshots(roll_for_evidence, str(tid))
        evidence_items = []
        for idx, v in enumerate(raw_violations, 1):
            match = _match_screenshot_for_violation(v, evidence_paths)
            if match is not None and match.exists():
                evidence_items.append((idx, v, match))

        if evidence_items:
            story.append(Spacer(1, 18))
            story.append(PageBreak())
            story.append(Paragraph(
                f"Visual Evidence ({len(evidence_items)} captures)",
                styles["Heading2"]))
            story.append(Paragraph(
                "Screenshots are listed in the same order as the violations table above.",
                styles["Italic"]))
            story.append(Spacer(1, 10))

            evidence_caption_style = ParagraphStyle(
                "EvidenceCaption", parent=styles["Normal"],
                fontSize=9, leading=12, spaceAfter=4,
            )

            for idx, v, img_path in evidence_items:
                ts_str = fmt_ist(v.get("created_at", ""))
                sev = v["severity"].upper()
                sev_color = "#c0392b" if v["severity"] == "high" else "#d68910"
                vtype_pretty = v["violation_type"].replace("_", " ").title()
                caption = (
                    f'<b>#{idx} — {vtype_pretty}</b>  ·  '
                    f'<font color="{sev_color}"><b>{sev}</b></font>  ·  '
                    f'{ts_str}'
                )
                detail = clean_details(v.get("details"))
                if detail:
                    caption += f'<br/><font size="8" color="#666">{detail}</font>'

                try:
                    img_flowable = Image(
                        str(img_path),
                        width=4.5 * inch, height=3.4 * inch,
                        kind="proportional",
                    )
                    story.append(KeepTogether([
                        Paragraph(caption, evidence_caption_style),
                        img_flowable,
                        Spacer(1, 14),
                    ]))
                except Exception as img_err:
                    story.append(Paragraph(
                        caption + f'  <font color="#999">(image unreadable: {img_err})</font>',
                        evidence_caption_style))
                    story.append(Spacer(1, 8))

        story.append(Spacer(1, 20))
        story.append(Paragraph("Answer Sheet", styles["Heading2"]))
        story.append(Spacer(1, 8))

        try:
            pdf_questions = _load_questions(teacher_id=tid)
            q_correct = {q["id"]: q["correct"] for q in pdf_questions}
            q_texts = {q["id"]: q.get("question", "")[:50] for q in pdf_questions}
        except Exception:
            q_correct = {}
            q_texts = {}

        if answers:
            ad = [["#", "Question", "Student", "Correct", "Result"]]
            for a in answers:
                qid = str(a["question_id"])
                correct = q_correct.get(qid, "?")
                is_right = str(a["answer"]) == correct
                ad.append([
                    f"Q{qid}",
                    q_texts.get(qid, "")[:40],
                    a["answer"],
                    correct,
                    "✓" if is_right else "✗",
                ])
            at = Table(ad, colWidths=[30, 180, 50, 50, 40])
            at.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
                ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,-1), 9),
                ("ROWBACKGROUNDS", (0,1), (-1,-1),
                 [colors.HexColor("#f8f9fa"), colors.white]),
                ("GRID",    (0,0), (-1,-1), 0.5, colors.grey),
                ("PADDING", (0,0), (-1,-1), 6),
            ]))
            story.append(at)

        story.append(Spacer(1, 20))
        story.append(Paragraph(
            f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')} IST | "
            f"Session: {session_id[:20]}...",
            styles["Normal"]))

        doc.build(story)
        buf.seek(0)
        fname = (f"report_{_safe_filename(exam.get('roll_number'), 'unknown')}_"
                 f"{now_ist().strftime('%Y%m%d')}.pdf")
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}"})

    except HTTPException:
        raise
    except Exception as e:
        print(f"[PDF] {e}")
        raise HTTPException(status_code=500, detail=f"PDF error: {e}")
# ─── 13. SCORECARD PDF ──────────────────────────────────

@router.get("/api/v1/admin/scorecard-pdf/{session_id:path}")
def scorecard_pdf(session_id: str, request: Request):
    """Generate a student-facing scorecard PDF."""
    teacher = require_admin(request)
    tid = teacher["id"]
    try:
        pdf_bytes, fname, _ = _build_scorecard_pdf(session_id, tid)
        return StreamingResponse(
            io.BytesIO(pdf_bytes), media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}"})
    except HTTPException:
        raise
    except Exception as e:
        print(f"[Scorecard PDF] {e}")
        raise HTTPException(status_code=500, detail=f"Scorecard PDF error: {e}")


# ─── 14. SCORECARD ZIP ──────────────────────────────────

@router.get("/api/v1/admin/scorecard-zip")
def scorecard_zip(request: Request, exam_id: str = None):
    """Generate a ZIP of all student scorecards for an exam."""
    teacher = require_admin(request)
    tid = teacher["id"]
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                         TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet

        sess_q = supabase.table("exam_sessions")\
            .select("session_key,roll_number,full_name,score,total,percentage,time_taken_secs,risk_score,started_at,submitted_at,exam_id")\
            .eq("status", SessionStatus.COMPLETED).eq("teacher_id", str(tid))
        if exam_id:
            sess_q = sess_q.eq("exam_id", exam_id)
        sessions = (sess_q.execute()).data or []
        if not sessions:
            raise HTTPException(status_code=404, detail="No completed sessions found")

        eid = exam_id or (sessions[0].get("exam_id") if sessions else None)
        questions = _load_questions(teacher_id=tid, exam_id=eid)
        config = None
        try:
            config = _load_exam_config(str(tid), exam_id=eid)
        except Exception:
            pass
        exam_title = (config or {}).get("title", "Exam")

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for sess in sessions:
                sid = sess["session_key"]
                ans_rows = (supabase.table("answers").select("question_id,answer")
                            .eq("session_key", sid)
                            .eq("teacher_id", str(tid)).execute()).data or []
                ans_map = {str(a["question_id"]): a["answer"] for a in ans_rows}

                pdf_buf = io.BytesIO()
                doc = SimpleDocTemplate(pdf_buf, pagesize=A4, topMargin=40, bottomMargin=40)
                styles = getSampleStyleSheet()
                story = []

                story.append(Paragraph(f"Scorecard — {exam_title}", styles["Title"]))
                story.append(Spacer(1, 12))

                score = sess.get("score", 0)
                total = sess.get("total", 0)
                pct = sess.get("percentage", 0)
                passed = pct >= 40

                info = [
                    ["Field", "Value"],
                    ["Student Name", sess.get("full_name", "")],
                    ["Roll Number", sess.get("roll_number", "")],
                    ["Date", fmt_ist(sess.get("submitted_at", sess.get("started_at", "")))],
                    ["Score", f"{score}/{total}"],
                    ["Percentage", f"{pct}%"],
                    ["Result", "PASS" if passed else "FAIL"],
                    ["Time Taken", f"{sess.get('time_taken_secs', 0) // 60}m {sess.get('time_taken_secs', 0) % 60}s"],
                ]
                t = Table(info, colWidths=[140, 330])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
                    ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                    ("FONTNAME",  (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",  (0,0), (-1, -1), 10),
                    ("ROWBACKGROUNDS", (0,1), (-1, -1),
                     [colors.HexColor("#f0f4ff"), colors.white]),
                    ("GRID",    (0,0), (-1, -1), 0.5, colors.grey),
                    ("PADDING", (0,0), (-1, -1), 8),
                ]))
                story.append(t)
                story.append(Spacer(1, 20))

                story.append(Paragraph("Question-wise Results", styles["Heading2"]))
                story.append(Spacer(1, 8))

                if questions:
                    qd = [["#", "Question", "Your Answer", "Correct", "Result"]]
                    for i, q in enumerate(questions, 1):
                        qid = str(q.get("question_id", q.get("id", "")))
                        correct_ans = str(q.get("correct", ""))
                        student_ans = ans_map.get(qid, "—")
                        is_right = str(student_ans) == correct_ans
                        q_text = q.get("question", "")
                        if len(q_text) > 60:
                            q_text = q_text[:57] + "..."
                        qd.append([str(i), q_text, str(student_ans)[:20], correct_ans[:20],
                                   "✓" if is_right else "✗"])
                    qt = Table(qd, colWidths=[25, 230, 80, 80, 35])
                    qt.setStyle(TableStyle([
                        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a1a2e")),
                        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                        ("FONTNAME",  (0,0), (-1,0), "Helvetica-Bold"),
                        ("FONTSIZE",  (0,0), (-1, -1), 9),
                        ("ROWBACKGROUNDS", (0,1), (-1, -1),
                         [colors.HexColor("#f8f9fa"), colors.white]),
                        ("GRID",    (0,0), (-1, -1), 0.5, colors.grey),
                        ("PADDING", (0,0), (-1, -1), 6),
                        ("ALIGN", (4,1), (4, -1), "CENTER"),
                    ]))
                    story.append(qt)

                story.append(Spacer(1, 20))
                story.append(Paragraph(
                    f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')} IST",
                    styles["Normal"]))

                doc.build(story)
                pdf_buf.seek(0)
                roll = _safe_filename(sess.get("roll_number"), "unknown")
                zf.writestr(f"scorecard_{roll}.pdf", pdf_buf.getvalue())

        zip_buf.seek(0)
        fname = f"scorecards_{_safe_filename(exam_id, 'all')}_{now_ist().strftime('%Y%m%d')}.zip"
        return StreamingResponse(
            zip_buf, media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={fname}"})

    except HTTPException:
        raise
    except Exception as e:
        print(f"[Scorecard ZIP] {e}")
        raise HTTPException(status_code=500, detail=f"Scorecard ZIP error: {e}")


# ─── 15. EMAIL SCORECARDS ────────────────────────────────

@router.post("/api/v1/admin/exams/{exam_id}/email-scorecards")
def email_scorecards(exam_id: str, request: Request, body: dict = Body(default={})):
    """Email every completed student their scorecard PDF for this exam."""
    from emailer import send_scorecard_email
    teacher = require_admin(request)
    tid = str(teacher["id"])

    resend_all = bool(body.get("resend_all"))
    custom_message = (body.get("custom_message") or "").strip() or None
    teacher_name = teacher.get("full_name") or teacher.get("email") or "Your teacher"

    sess_q = (supabase.table("exam_sessions").select(
        "session_key,roll_number,full_name,exam_id,scorecard_emailed_at"
    ).eq("teacher_id", tid).eq("status", SessionStatus.COMPLETED).eq("exam_id", exam_id)
        .limit(1000))
    sessions = (sess_q.execute()).data or []
    if not sessions:
        raise HTTPException(status_code=404, detail="No completed sessions found for this exam")

    roll_emails: dict[str, str] = {}
    try:
        inv_rows = (supabase.table("student_invites").select("roll_number,email")
                    .eq("teacher_id", tid).eq("exam_id", exam_id).execute()).data or []
        for r in inv_rows:
            roll = str(r.get("roll_number") or "").strip().upper()
            email = str(r.get("email") or "").strip().lower()
            if roll and email:
                roll_emails[roll] = email
    except Exception as e:
        print(f"[email-scorecards] invite lookup failed: {e}", flush=True)
    try:
        stud_rows = (supabase.table("students").select("roll_number,email")
                     .eq("teacher_id", tid).execute()).data or []
        for r in stud_rows:
            roll = str(r.get("roll_number") or "").strip().upper()
            email = str(r.get("email") or "").strip().lower()
            if roll and email and roll not in roll_emails:
                roll_emails[roll] = email
    except Exception as e:
        print(f"[email-scorecards] student lookup failed: {e}", flush=True)

    sent = 0
    failed = 0
    already_sent = 0
    skipped_no_email = 0
    failures: list[dict] = []

    for sess in sessions:
        sid = sess["session_key"]
        roll = str(sess.get("roll_number") or "").strip().upper()
        full_name = sess.get("full_name") or "Student"

        if sess.get("scorecard_emailed_at") and not resend_all:
            already_sent += 1
            continue

        email = roll_emails.get(roll)
        if not email:
            skipped_no_email += 1
            failures.append({"roll": roll, "reason": "no email on file"})
            continue

        now_iso = datetime.now(timezone.utc).isoformat()
        if not resend_all:
            claim = (supabase.table("exam_sessions")
                     .update({"scorecard_emailed_at": now_iso})
                     .eq("session_key", sid)
                     .eq("teacher_id", tid)
                     .is_("scorecard_emailed_at", "null")
                     .execute())
            if not claim.data:
                already_sent += 1
                continue

        try:
            pdf_bytes, fname, summary = _build_scorecard_pdf(sid, tid)
        except Exception as e:
            print(f"[email-scorecards] PDF build failed sid={sid} err={e}", flush=True)
            if not resend_all:
                try:
                    (supabase.table("exam_sessions")
                     .update({"scorecard_emailed_at": None})
                     .eq("session_key", sid).eq("teacher_id", tid).execute())
                except Exception:
                    pass
            failed += 1
            failures.append({"roll": roll, "reason": f"pdf: {e}"})
            continue

        result = send_scorecard_email(
            to_email=email,
            to_name=full_name,
            exam_title=summary.get("exam_title") or "Exam",
            score=int(summary.get("score") or 0),
            total=int(summary.get("total") or 0),
            percentage=float(summary.get("percentage") or 0.0),
            passed=bool(summary.get("passed")),
            pdf_bytes=pdf_bytes,
            pdf_filename=fname,
            teacher_name=teacher_name,
            custom_message=custom_message,
        )

        if result.ok:
            try:
                update_row = {"scorecard_email_msg_id": result.provider_msg_id}
                if resend_all:
                    update_row["scorecard_emailed_at"] = now_iso
                (supabase.table("exam_sessions").update(update_row)
                 .eq("session_key", sid).eq("teacher_id", tid).execute())
            except Exception as e:
                print(f"[email-scorecards] msg_id update failed sid={sid}: {e}", flush=True)
            sent += 1
        else:
            if not resend_all:
                try:
                    (supabase.table("exam_sessions")
                     .update({"scorecard_emailed_at": None})
                     .eq("session_key", sid).eq("teacher_id", tid).execute())
                except Exception:
                    pass
            failed += 1
            failures.append({"roll": roll, "reason": result.error or "send failed"})
            print(f"[email-scorecards][SEND_ERROR] roll={roll} reason={result.error!r}", flush=True)

    return {
        "sent": sent,
        "failed": failed,
        "already_sent": already_sent,
        "skipped_no_email": skipped_no_email,
        "total": len(sessions),
        "failures": failures[:50],
    }
# ─── 16. FAILED SESSIONS ────────────────────────────────#

@router.get("/api/v1/admin-failed-sessions")
def failed_sessions(request: Request, exam_id: str = None):
    """Returns sessions with submit_failed events that never completed."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    failed = supabase.table("violations").select("session_key")\
        .eq("violation_type", "submit_failed")\
        .eq("teacher_id", tid)\
        .execute()
    failed_keys = {r["session_key"] for r in (failed.data or [])}
    sub_query = supabase.table("exam_sessions").select("session_key")\
        .eq("status", SessionStatus.COMPLETED)\
        .eq("teacher_id", tid)\
        .in_("session_key", list(failed_keys) or ["__none__"])
    if exam_id:
        sub_query = sub_query.eq("exam_id", exam_id)
    submitted = sub_query.execute()
    submitted_keys = {r["session_key"] for r in (submitted.data or [])}
    if exam_id:
        es = supabase.table("exam_sessions").select("session_key")\
            .eq("teacher_id", tid).eq("exam_id", exam_id).execute()
        exam_skeys = {r["session_key"] for r in (es.data or [])}
        failed_keys = failed_keys & exam_skeys
    unrecovered = [k for k in failed_keys if k not in submitted_keys]
    return {"failed_sessions": unrecovered, "count": len(unrecovered)}


# ─── 17. CLEANUP SCREENSHOTS ────────────────────────────#

@router.post("/api/v1/admin-cleanup")
def admin_cleanup(request: Request):
    """Delete the calling teacher's screenshots older than 7 days."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    deleted = 0
    cutoff  = now_ist() - timedelta(days=7)
    teacher_root = Path(SCREENSHOTS_DIR) / tid
    if not teacher_root.is_dir():
        return {"deleted": 0}
    try:
        for student_dir in teacher_root.iterdir():
            if student_dir.is_dir():
                for f in student_dir.iterdir():
                    if f.is_file() and f.stat().st_mtime < cutoff.timestamp():
                        f.unlink()
                        deleted += 1
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"deleted": deleted}


# ─── 18. CLEAR LIVE SESSIONS ─────────────────────────────#

@router.post("/api/v1/admin/clear-live-sessions")
def clear_live_sessions(request: Request, body: dict = Body(...)):
    """Destructive: wipe all in-progress sessions for the calling teacher."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    step = str(body.get("step") or "").lower().strip()

    include_completed = bool(body.get("include_completed", False))
    include_active = bool(body.get("include_active", False))
    raw_eid = body.get("exam_id") or None
    exam_id_scope: str | None = str(raw_eid).strip() or None if raw_eid else None

    if step == "request":
        active, stale = _partition_live_sessions(
            tid, exam_id=exam_id_scope, include_active=include_active,
        )
        completed_rows: list[dict] = []
        if include_completed:
            comp_q = supabase.table("exam_sessions")\
                .select("session_key,roll_number,full_name,started_at,submitted_at,exam_id")\
                .eq("teacher_id", tid)\
                .eq("status", SessionStatus.COMPLETED)
            if exam_id_scope:
                comp_q = comp_q.eq("exam_id", exam_id_scope)
            comp = comp_q.execute()
            completed_rows = comp.data or []
        token = _clear_token_issue(tid)
        return {
            "step":          "request",
            "token":          token,
            "expires_in":     _CLEAR_TOKEN_TTL,
            "active_window_s": _CLEAR_ACTIVE_WINDOW,
            "include_completed": include_completed,
            "include_active":    include_active,
            "exam_id":           exam_id_scope or "",
            "count":          len(stale) + len(completed_rows),
            "stale_count":    len(stale),
            "active_count":   len(active),
            "completed_count": len(completed_rows),
            "preview":    [
                {"session_key": r["session_key"],
                 "roll_number": r.get("roll_number"),
                 "full_name":   r.get("full_name"),
                 "started_at":  r.get("started_at"),
                 "last_heartbeat": r.get("last_heartbeat")}
                for r in stale[:20]
            ],
            "active_preview": [
                {"session_key": r["session_key"],
                 "roll_number": r.get("roll_number"),
                 "full_name":   r.get("full_name"),
                 "last_heartbeat": r.get("last_heartbeat")}
                for r in active[:20]
            ],
            "completed_preview": [
                {"session_key": r["session_key"],
                 "roll_number": r.get("roll_number"),
                 "full_name":   r.get("full_name"),
                 "submitted_at": r.get("submitted_at")}
                for r in completed_rows[:20]
            ],
        }

    if step == "confirm":
        token = str(body.get("token") or "")
        ack   = str(body.get("ack") or "")
        if ack != "DELETE":
            raise HTTPException(status_code=400,
                detail="Missing or incorrect ack — expected 'DELETE'")
        if not _clear_token_consume(token, tid):
            raise HTTPException(status_code=400,
                detail="Confirmation token is invalid or expired — restart the clear flow")

        active, stale = _partition_live_sessions(
            tid, exam_id=exam_id_scope, include_active=include_active,
        )

        completed_keys: list[str] = []
        comp = None
        if include_completed:
            comp_q = supabase.table("exam_sessions")\
                .select("session_key,roll_number,exam_id")\
                .eq("teacher_id", tid)\
                .eq("status", SessionStatus.COMPLETED)
            if exam_id_scope:
                comp_q = comp_q.eq("exam_id", exam_id_scope)
            comp = comp_q.execute()
            completed_keys = [r["session_key"] for r in (comp.data or [])]

        if not stale and not completed_keys:
            skipped_active = [
                {"session_key": r["session_key"],
                 "roll_number": r.get("roll_number"),
                 "full_name":   r.get("full_name")}
                for r in active
            ]
            return {"step": "confirm", "cleared": 0, "sessions": 0,
                    "answers": 0, "violations": 0, "screenshots": 0,
                    "skipped_active": len(active), "skipped": skipped_active,
                    "note": ("No sessions to clear"
                             + (" — active students were protected"
                                if active else ""))}

        session_keys = [r["session_key"] for r in stale] + completed_keys
        rolls_seen = set()
        for r in stale:
            if r.get("roll_number"):
                rolls_seen.add(r["roll_number"])
        if include_completed:
            for r in (comp.data or []):
                if r.get("roll_number"):
                    rolls_seen.add(r["roll_number"])

        skipped_active = [
            {"session_key": r["session_key"],
             "roll_number": r.get("roll_number"),
             "full_name":   r.get("full_name")}
            for r in active
        ]
        if active:
            print(f"[ClearLive] teacher={tid} protecting {len(active)} "
                  f"active session(s) from wipe")

        ans_deleted = 0
        viol_deleted = 0
        scr_deleted = 0

        _sk_tid = {r["session_key"]: r.get("teacher_id") or ""
                   for r in stale}
        _ghost_keys = {r["session_key"] for r in stale if r.get("_ghost")}

        for sk in session_keys:
            sk_tid = _sk_tid.get(sk, tid)
            is_ghost = sk in _ghost_keys
            try:
                q = supabase.table("answers").delete().eq("session_key", sk)
                if sk_tid and not is_ghost:
                    q = q.eq("teacher_id", sk_tid)
                r = q.execute()
                ans_deleted += len(r.data or [])
            except Exception as e:
                print(f"[ClearLive] answer delete failed {sk}: {e}")
            try:
                q = supabase.table("violations").delete().eq("session_key", sk)
                if sk_tid and not is_ghost:
                    q = q.eq("teacher_id", sk_tid)
                r = q.execute()
                viol_deleted += len(r.data or [])
            except Exception as e:
                print(f"[ClearLive] violation delete failed {sk}: {e}")

        stale_key_set = {r["session_key"] for r in stale}
        sess_deleted = 0
        for sk in session_keys:
            try:
                if sk in _ghost_keys:
                    supabase.table("exam_sessions").delete()\
                        .eq("session_key", sk).execute()
                else:
                    q = supabase.table("exam_sessions").delete()\
                        .eq("session_key", sk)
                    sk_tid = _sk_tid.get(sk, tid)
                    if sk_tid:
                        q = q.eq("teacher_id", sk_tid)
                    if sk in stale_key_set:
                        q = q.eq("status", SessionStatus.IN_PROGRESS)
                    else:
                        q = q.eq("status", SessionStatus.COMPLETED)
                    q.execute()
                sess_deleted += 1
            except Exception as e:
                print(f"[ClearLive] session delete failed {sk}: {e}")

        active_rolls = {r.get("roll_number") for r in active if r.get("roll_number")}
        t_screens = Path(SCREENSHOTS_DIR) / tid
        if t_screens.is_dir():
            for roll in rolls_seen:
                if not roll:
                    continue
                if roll in active_rolls:
                    continue
                safe = _safe_path_component(roll)
                rdir = t_screens / safe
                if not rdir.is_dir():
                    continue
                if not include_completed:
                    comp_chk = supabase.table("exam_sessions")\
                        .select("session_key", count="exact")\
                        .eq("teacher_id", tid)\
                        .eq("roll_number", roll)\
                        .eq("status", SessionStatus.COMPLETED)\
                        .execute()
                    if (comp_chk.count or 0) > 0:
                        continue
                try:
                    for f in rdir.iterdir():
                        if f.is_file():
                            f.unlink()
                            scr_deleted += 1
                    rdir.rmdir()
                except Exception as e:
                    print(f"[ClearLive] screenshot cleanup failed {rdir}: {e}")

        print(f"[ClearLive] teacher={tid} sessions={sess_deleted} "
              f"(completed={len(completed_keys)}) "
              f"answers={ans_deleted} violations={viol_deleted} "
              f"screenshots={scr_deleted} "
              f"protected_active={len(active)}")
        return {
            "step":           "confirm",
            "cleared":        sess_deleted,
            "sessions":       sess_deleted,
            "answers":        ans_deleted,
            "violations":     viol_deleted,
            "screenshots":    scr_deleted,
            "completed_cleared": len(completed_keys),
            "skipped_active": len(active),
            "skipped":        skipped_active,
        }

    raise HTTPException(status_code=400,
        detail="'step' must be 'request' or 'confirm'")


# ─── 19. BACKFILL RISK SCORES ──────────────────────────────#

@router.post("/api/v1/admin/backfill-risk-scores")
def backfill_risk_scores(request: Request, exam_id: str = None):
    """Recompute and cache risk scores for all completed sessions."""
    teacher = require_admin(request)
    tid = teacher["id"]
    query = supabase.table("exam_sessions").select("session_key")\
        .eq("status", SessionStatus.COMPLETED)\
        .eq("teacher_id", str(tid))
    if exam_id:
        query = query.eq("exam_id", exam_id)
    sessions = query.execute()
    count = 0
    for s in (sessions.data or []):
        risk = compute_risk_score(s["session_key"], teacher_id=tid)
        supabase.table("exam_sessions").update(
            {"risk_score": risk["risk_score"]}
        ).eq("session_key", s["session_key"])\
             .eq("teacher_id", str(tid))\
             .execute()
        count += 1
    return {"backfilled": count}
# ─── 20. LIST EXAMS ─────────────────────────────────#

@router.get("/api/v1/admin/exams")
def list_exams(request: Request):
    """List all exams for the calling teacher."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    result = supabase.table("exam_config").select("*").eq("teacher_id", tid).execute()
    exams = result.data or []
    out = []
    for ex in exams:
        eid = ex.get("exam_id")
        qcount = 0
        scount = 0
        try:
            qr = supabase.table("questions").select("question_id", count="exact")\
                .eq("teacher_id", tid).eq("exam_id", eid).execute()
            qcount = qr.count if qr.count is not None else len(qr.data or [])
        except Exception:
            pass
        try:
            sr = supabase.table("exam_sessions").select("session_key", count="exact")\
                .eq("teacher_id", tid).eq("exam_id", eid).execute()
            scount = sr.count if sr.count is not None else len(sr.data or [])
        except Exception:
            pass
        out.append({
            "exam_id":          eid,
            "exam_title":       ex.get("exam_title", "Exam"),
            "duration_minutes": ex.get("duration_minutes", 60),
            "starts_at":        ex.get("starts_at"),
            "ends_at":          ex.get("ends_at"),
            "access_code":      ex.get("access_code", ""),
            "question_count":   qcount,
            "session_count":    scount,
            "created_at":       ex.get("created_at", ""),
        })
    return {"exams": out}


# ─── 21. CREATE EXAM ─────────────────────────────────#

@router.post("/api/v1/admin/exams")
def create_exam(request: Request, body: dict = Body(...)):
    """Create a new exam for the calling teacher."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    title = str(body.get("exam_title", "New Exam")).strip() or "New Exam"
    duration = int(body.get("duration_minutes", 60))
    exam_id = str(_uuid.uuid4())
    try:
        result = supabase.table("exam_config").insert({
            "exam_id":          exam_id,
            "teacher_id":       tid,
            "exam_title":       title,
            "duration_minutes": duration,
        }).execute()
    except Exception as e:
        print(f"[CreateExam] DB error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create exam: {e}")
    row = result.data[0] if result.data else {}
    return {"exam_id": row.get("exam_id", exam_id), "exam_title": title, "duration_minutes": duration}


# ─── 22. DELETE EXAM ─────────────────────────────────#

@router.delete("/api/v1/admin/exams/{exam_id}")
def delete_exam(exam_id: str, request: Request):
    """Delete an exam and its questions. Keeps session history."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    check = supabase.table("exam_config").select("exam_id")\
        .eq("teacher_id", tid).eq("exam_id", exam_id).execute()
    if not check.data:
        raise HTTPException(status_code=404, detail="Exam not found")
    all_exams = supabase.table("exam_config").select("exam_id")\
        .eq("teacher_id", tid).execute()
    if len(all_exams.data or []) <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete your only exam")
    supabase.table("questions").delete()\
        .eq("teacher_id", tid).eq("exam_id", exam_id).execute()
    supabase.table("exam_config").delete()\
        .eq("teacher_id", tid).eq("exam_id", exam_id).execute()
    if _cache:
        _cache.delete(f"exam_config:{tid}:{exam_id or '_'}")
        _cache.delete(f"questions:{tid}:{exam_id or '_'}")
    return {"status": "deleted", "exam_id": exam_id}


# ─── 23. DUPLICATE EXAM ──────────────────────────────#

@router.post("/api/v1/admin/exams/{exam_id}/duplicate")
def duplicate_exam(exam_id: str, request: Request, body: dict = Body(default={})):
    """Clone an exam's config + questions into a fresh exam_id."""
    teacher = require_admin(request)
    tid = str(teacher["id"])

    src_q = (supabase.table("exam_config").select("*")
             .eq("teacher_id", tid).eq("exam_id", exam_id).execute())
    if not src_q.data:
        raise HTTPException(status_code=404, detail="Exam not found")
    src = src_q.data[0]

    new_exam_id = str(_uuid.uuid4())
    src_title = src.get("exam_title") or "Exam"
    new_title = (str(body.get("new_title") or "").strip()
                 or f"{src_title} (copy)")

    COPYABLE = [
        "duration_minutes",
        "shuffle_questions", "shuffle_options",
    ]
    new_cfg = {
        "exam_id":    new_exam_id,
        "teacher_id": tid,
        "exam_title": new_title,
        "starts_at":  None,
        "ends_at":    None,
        "access_code": "",
    }
    for col in COPYABLE:
        if col in src and src[col] is not None:
            new_cfg[col] = src[col]

    try:
        supabase.table("exam_config").insert(new_cfg).execute()
    except Exception as e:
        print(f"[DuplicateExam] config insert failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to clone config: {e}")

    try:
        qsrc = (supabase.table("questions").select("*")
                .eq("teacher_id", tid).eq("exam_id", exam_id)
                .order("order_index").execute()).data or []
    except Exception as e:
        print(f"[DuplicateExam] question fetch failed: {e}")
        qsrc = []

    questions_copied = 0
    if qsrc:
        new_rows = []
        for q in qsrc:
            row = dict(q)
            for k in ("id", "question_id", "created_at", "updated_at"):
                row.pop(k, None)
            row["exam_id"] = new_exam_id
            row["teacher_id"] = tid
            new_rows.append(row)
        try:
            for i in range(0, len(new_rows), 500):
                supabase.table("questions").insert(new_rows[i:i+500]).execute()
                questions_copied += len(new_rows[i:i+500])
        except Exception as e:
            print(f"[DuplicateExam] question insert failed: {e}")
            try:
                supabase.table("exam_config").delete()\
                    .eq("teacher_id", tid).eq("exam_id", new_exam_id).execute()
                supabase.table("questions").delete()\
                    .eq("teacher_id", tid).eq("exam_id", new_exam_id).execute()
            except Exception:
                pass
            raise HTTPException(status_code=500, detail=f"Failed to clone questions: {e}")

    if _cache:
        _cache.delete(f"exam_config:{tid}:{new_exam_id}")
        _cache.delete(f"questions:{tid}:{new_exam_id}")

    return {
        "status":           "duplicated",
        "source_exam_id":   exam_id,
        "exam_id":          new_exam_id,
        "exam_title":       new_title,
        "questions_copied": questions_copied,
    }
# ─── 24. ANALYTICS ─────────────────────────────────#

@router.get("/api/v1/admin/analytics")
def get_analytics(request: Request):
    """Compute exam analytics: score distribution, question analysis, violations, risk."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    exam_id = request.query_params.get("exam_id")

    cache_key = f"analytics:{tid}:{exam_id or '_'}"
    if _cache:
        cached = _cache.get(cache_key)
        if cached:
            return cached

    sess_q = supabase.table("exam_sessions")\
        .select("session_key,roll_number,full_name,score,total,percentage,time_taken_secs,risk_score,started_at")\
        .eq("status", SessionStatus.COMPLETED)
    if tid:
        sess_q = sess_q.eq("teacher_id", tid)
    if exam_id:
        sess_q = sess_q.eq("exam_id", exam_id)
    sessions = (sess_q.execute()).data or []

    if not sessions:
        empty = {"exam_overview": {"count": 0}, "score_distribution": [],
                 "question_analysis": [], "violation_summary": {},
                 "risk_distribution": {"low": 0, "medium": 0, "high": 0}}
        if _cache:
            _cache.set(cache_key, empty, ttl=60)
        return empty

    count = len(sessions)
    pcts = [s.get("percentage") or 0 for s in sessions]
    times = [s.get("time_taken_secs") or 0 for s in sessions]
    scores = [s.get("score") or 0 for s in sessions]
    totals = [s.get("total") or 1 for s in sessions]
    avg_score = round(sum(scores) / count, 1)
    avg_pct = round(sum(pcts) / count, 1)
    sorted_times = sorted(t for t in times if t > 0)
    median_time = sorted_times[len(sorted_times)//2] if sorted_times else 0
    pass_count = sum(1 for p in pcts if p >= 40)
    overview = {
        "count": count,
        "avg_score": avg_score,
        "avg_total": round(sum(totals) / count, 1),
        "avg_percentage": avg_pct,
        "median_time_secs": median_time,
        "pass_rate": round(pass_count / count * 100, 1),
    }

    buckets = [0] * 10
    for p in pcts:
        idx = min(int(p // 10), 9)
        buckets[idx] += 1
    score_dist = [{"range": f"{i*10}-{i*10+10}%", "count": buckets[i]} for i in range(10)]

    questions = _load_questions(tid, exam_id=exam_id)
    q_analysis = []
    if questions:
        skeys = [s["session_key"] for s in sessions]
        all_answers = {sk: {} for sk in skeys}
        for i in range(0, len(skeys), 50):
            chunk = skeys[i:i+50]
            ans_q = (supabase.table("answers")
                     .select("session_key,question_id,answer")
                     .in_("session_key", chunk))
            if tid:
                ans_q = ans_q.eq("teacher_id", tid)
            for r in (ans_q.execute()).data or []:
                sk = r.get("session_key")
                qid = r.get("question_id")
                if sk and qid is not None:
                    all_answers.setdefault(sk, {})[qid] = r.get("answer")

        sorted_sess = sorted(sessions, key=lambda s: s.get("percentage") or 0)
        q1_cutoff = max(1, count // 4)
        bottom_keys = set(s["session_key"] for s in sorted_sess[:q1_cutoff])
        top_keys = set(s["session_key"] for s in sorted_sess[-q1_cutoff:])

        for q in questions:
            qid = str(q.get("question_id") or q.get("id", ""))
            correct = str(q.get("correct", ""))
            total_attempted = 0
            total_correct = 0
            top_correct = 0
            top_total = 0
            bottom_correct = 0
            bottom_total = 0
            for sk, ans_map in all_answers.items():
                if qid in ans_map:
                    total_attempted += 1
                    is_correct = ans_map[qid] == correct
                    if is_correct:
                        total_correct += 1
                    if sk in top_keys:
                        top_total += 1
                        if is_correct:
                            top_correct += 1
                    if sk in bottom_keys:
                        bottom_total += 1
                        if is_correct:
                            bottom_correct += 1
            difficulty = round(total_correct / max(total_attempted, 1) * 100, 1)
            top_rate = top_correct / max(top_total, 1)
            bottom_rate = bottom_correct / max(bottom_total, 1)
            discrimination = round(top_rate - bottom_rate, 2)
            q_analysis.append({
                "question_id": qid,
                "question": (q.get("question", "")[:80] + "...") if len(q.get("question", "")) > 80 else q.get("question", ""),
                "difficulty_pct": difficulty,
                "discrimination": discrimination,
                "attempted": total_attempted,
                "correct": total_correct,
            })

    viol_q = supabase.table("violations")\
        .select("violation_type,severity,session_key,created_at")
    if tid:
        viol_q = viol_q.eq("teacher_id", tid)
    viols = (viol_q.execute()).data or []
    scored_viols = [v for v in viols if True and v.get("severity") in ("high", "medium")]

    type_counts = {}
    for v in scored_viols:
        vt = v["violation_type"]
        type_counts[vt] = type_counts.get(vt, 0) + 1
    viol_summary = {"by_type": type_counts, "total": len(scored_viols)}

    risk_dist = {"low": 0, "medium": 0, "high": 0}
    for s in sessions:
        rs = s.get("risk_score") or 0
        if rs <= 30:
            risk_dist["low"] += 1
        elif rs <= 60:
            risk_dist["medium"] += 1
        else:
            risk_dist["high"] += 1

    result = {
        "exam_overview": overview,
        "score_distribution": score_dist,
        "question_analysis": q_analysis,
        "violation_summary": viol_summary,
        "risk_distribution": risk_dist,
    }
    if _cache:
        _cache.set(cache_key, result, ttl=60)
    return result


# ─── 25. LIST GROUPS ─────────────────────────────────#

@router.get("/api/v1/admin/groups")
def list_groups(request: Request):
    """List all groups for the authenticated teacher."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    rows = (supabase.table("student_groups")
            .select("*").eq("teacher_id", tid)
            .order("created_at").execute()).data or []
    counts: dict[str, int] = {}
    if rows:
        gids = [g["id"] for g in rows]
        members = (supabase.table("student_group_members")
                   .select("group_id")
                   .in_("group_id", gids)
                   .eq("teacher_id", tid)
                   .limit(50000).execute()).data or []
        for m in members:
            gid = m.get("group_id")
            if gid:
                counts[gid] = counts.get(gid, 0) + 1
    for g in rows:
        g["member_count"] = counts.get(g["id"], 0)
    return rows
# ─── 26. CREATE GROUP ─────────────────────────────────#

@router.post("/api/v1/admin/groups")
def create_group(request: Request, body: dict = Body(...)):
    """Create a new student group."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    name = (body.get("group_name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="group_name is required")
    try:
        row = (supabase.table("student_groups")
               .insert({"teacher_id": tid, "group_name": name}).execute()).data
        return row[0] if row else {"ok": True}
    except Exception as e:
        if "duplicate" in str(e).lower() or "unique" in str(e).lower():
            raise HTTPException(status_code=409, detail="Group name already exists")
        raise


# ─── 27. RENAME GROUP ─────────────────────────────────#

@router.put("/api/v1/admin/groups/{group_id}")
def rename_group(group_id: str, request: Request, body: dict = Body(...)):
    """Rename a student group."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    name = (body.get("group_name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="group_name is required")
    result = (supabase.table("student_groups")
              .update({"group_name": name})
              .eq("id", group_id).eq("teacher_id", tid).execute())
    if not result.data:
        raise HTTPException(status_code=404, detail="Group not found")
    return result.data[0]


# ─── 28. DELETE GROUP ─────────────────────────────────#

@router.delete("/api/v1/admin/groups/{group_id}")
def delete_group(group_id: str, request: Request):
    """Delete a student group (cascades to members and exam assignments)."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    result = (supabase.table("student_groups")
              .delete().eq("id", group_id).eq("teacher_id", tid).execute())
    if not result.data:
        raise HTTPException(status_code=404, detail="Group not found")
    return {"ok": True}


# ─── 29. LIST GROUP MEMBERS ──────────────────────────#

@router.get("/api/v1/admin/groups/{group_id}/members")
def list_group_members(group_id: str, request: Request):
    """List members of a group, enriched with email/full_name."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    rows = (supabase.table("student_group_members")
            .select("*").eq("group_id", group_id)
            .eq("teacher_id", tid).execute()).data or []
    if not rows:
        return []
    rolls = [r["roll_number"] for r in rows if r.get("roll_number")]
    if rolls:
        students = (supabase.table("students")
                    .select("roll_number,email,full_name")
                    .eq("teacher_id", tid)
                    .in_("roll_number", rolls).execute()).data or []
        by_roll = {s["roll_number"]: s for s in students}
        for r in rows:
            s = by_roll.get(r.get("roll_number")) or {}
            r["email"] = s.get("email") or ""
            r["full_name"] = s.get("full_name") or ""
    return rows


# ─── 30. ADD GROUP MEMBERS ──────────────────────────────#

@router.post("/api/v1/admin/groups/{group_id}/members")
def add_group_members(group_id: str, request: Request, body: dict = Body(...)):
    """Add students to a group by roll numbers."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    grp = (supabase.table("student_groups")
           .select("id").eq("id", group_id).eq("teacher_id", tid).execute()).data
    if not grp:
        raise HTTPException(status_code=404, detail="Group not found")
    rolls = body.get("roll_numbers", [])
    if not rolls:
        raise HTTPException(status_code=400, detail="roll_numbers list is required")
    rows = [{"group_id": group_id, "roll_number": str(r).strip(), "teacher_id": tid}
            for r in rolls if str(r).strip()]
    if rows:
        supabase.table("student_group_members").upsert(rows).execute()
    return {"added": len(rows)}


# ─── 31. REMOVE GROUP MEMBERS ───────────────────────────#

@router.delete("/api/v1/admin/groups/{group_id}/members")
def remove_group_members(group_id: str, request: Request, body: dict = Body(...)):
    """Remove students from a group by roll numbers."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    rolls = body.get("roll_numbers", [])
    if not rolls:
        raise HTTPException(status_code=400, detail="roll_numbers list is required")
    for r in rolls:
        supabase.table("student_group_members")\
            .delete().eq("group_id", group_id)\
            .eq("roll_number", str(r).strip())\
            .eq("teacher_id", tid).execute()
    return {"removed": len(rolls)}


# ─── 32. LIST EXAM GROUPS ──────────────────────────────#

@router.get("/api/v1/admin/exams/{exam_id}/groups")
def list_exam_groups(exam_id: str, request: Request):
    """List groups assigned to an exam."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    assignments = (supabase.table("exam_group_assignments")
                   .select("group_id").eq("exam_id", exam_id)
                   .eq("teacher_id", tid).execute()).data or []
    if not assignments:
        return []
    gids = [a["group_id"] for a in assignments]
    groups = (supabase.table("student_groups")
              .select("*").in_("id", gids).execute()).data or []
    return groups


# ─── 33. ASSIGN GROUPS TO EXAM ───────────────────────────#

@router.post("/api/v1/admin/exams/{exam_id}/groups")
def assign_exam_groups(exam_id: str, request: Request, body: dict = Body(...)):
    """Assign groups to an exam for access control."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    group_ids = body.get("group_ids", [])
    if not group_ids:
        raise HTTPException(status_code=400, detail="group_ids list is required")
    rows = [{"exam_id": exam_id, "group_id": gid, "teacher_id": tid} for gid in group_ids]
    supabase.table("exam_group_assignments").upsert(rows).execute()
    return {"assigned": len(rows)}


# ─── 34. UNASSIGN GROUP FROM EXAM ────────────────────────#

@router.delete("/api/v1/admin/exams/{exam_id}/groups/{group_id}")
def unassign_exam_group(exam_id: str, group_id: str, request: Request):
    """Remove a group assignment from an exam."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    supabase.table("exam_group_assignments")\
        .delete().eq("exam_id", exam_id)\
        .eq("group_id", group_id)\
        .eq("teacher_id", tid).execute()
    return {"ok": True}


# ─── 35. BULK REGISTER STUDENTS ───────────────────────────#

@router.post("/api/v1/admin/register-students-bulk")
def admin_bulk_register(request: Request, body: dict = Body(...)):
    """Admin-only bulk student registration."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    students = body.get("students", [])
    if not students or not isinstance(students, list):
        raise HTTPException(status_code=400, detail="'students' must be a non-empty list")
    if len(students) > 500:
        raise HTTPException(status_code=400, detail="Max 500 students per batch")

    rows = []
    for s in students:
        roll = str(s.get("roll_number", "")).strip().upper()
        name = str(s.get("full_name", "")).strip()
        email = str(s.get("email", "")).strip().lower()
        phone = str(s.get("phone", "")).strip() or None
        if not roll or not name or not email:
            continue
        rows.append({
            "roll_number": roll,
            "full_name": name,
            "email": email,
            "phone": phone,
            "teacher_id": tid,
        })

    if not rows:
        raise HTTPException(status_code=400, detail="No valid students in payload")

    registered = 0
    skipped = 0
    for row in rows:
        try:
            supabase.table("students").insert(row).execute()
            registered += 1
        except Exception as e:
            if "duplicate" in str(e).lower() or "unique" in str(e).lower():
                skipped += 1
            else:
                skipped += 1

    return {"registered": registered, "skipped": skipped, "total": len(rows)}
# ─── 36. GET ACCESS CODE ─────────────────────────#

@router.get("/api/v1/admin/access-code")
def get_access_code(request: Request):
    """Return the current exam access code."""
    teacher = require_admin(request)
    exam_id = request.query_params.get("exam_id")
    code = _get_access_code(teacher["id"], exam_id=exam_id)
    return {"access_code": code, "enabled": bool(code)}


# ─── 37. SET ACCESS CODE ─────────────────────────#

@router.post("/api/v1/admin/access-code")
def set_access_code(request: Request, body: dict = Body(...)):
    """Set or clear the exam access code."""
    teacher = require_admin(request)
    exam_id = body.get("exam_id")
    new_code = str(body.get("access_code", "")).strip().upper()
    _set_access_code(new_code, teacher["id"], exam_id=exam_id)
    if _cache:
        _cache.delete(f"exam_config:{teacher['id']}:{exam_id or '_'}")
    return {"access_code": new_code, "enabled": bool(new_code)}


# ─── 38. REGISTERED COUNT ────────────────────────#

@router.get("/api/v1/admin/registered-count")
def registered_count(request: Request):
    """Return total number of registered students."""
    teacher = require_admin(request)
    tid = teacher["id"]
    query = supabase.table("students").select("roll_number", count="exact")
    if tid:
        query = query.eq("teacher_id", tid)
    result = query.execute()
    return {"count": result.count if result.count is not None else len(result.data or [])}


# ─── 39. GET EXAM SCHEDULE ────────────────────────#

@router.get("/api/v1/admin/exam-schedule")
def admin_get_schedule(request: Request):
    """Return current exam schedule for the admin dashboard."""
    teacher = require_admin(request)
    exam_id = request.query_params.get("exam_id")
    config = _load_exam_config(teacher["id"], exam_id=exam_id)
    return {
        "exam_title": config.get("exam_title", "Exam"),
        "starts_at":  config.get("starts_at"),
        "ends_at":    config.get("ends_at"),
    }


# ─── 40. SET EXAM SCHEDULE ────────────────────────#

@router.post("/api/v1/admin/exam-schedule")
def admin_set_schedule(request: Request, body: dict = Body(...)):
    """Set or clear exam start/end times."""
    teacher = require_admin(request)
    tid = teacher["id"]
    exam_id = body.get("exam_id")

    if not exam_id:
        raise HTTPException(status_code=400, detail="exam_id is required")

    update = {}
    if "starts_at" in body:
        update["starts_at"] = body["starts_at"]
    if "ends_at" in body:
        update["ends_at"] = body["ends_at"]
    if update:
        supabase.table("exam_config").update(update)\
            .eq("teacher_id", tid).eq("exam_id", exam_id).execute()

    if _cache:
        _cache.delete(f"exam_config:{tid}:{exam_id}")
    return {
        "status":    "updated",
        "starts_at": body.get("starts_at"),
        "ends_at":   body.get("ends_at"),
    }


# ─── 41. GET SHUFFLE CONFIG ────────────────────────#

@router.get("/api/v1/admin/shuffle-config")
def admin_get_shuffle(request: Request):
    """Return current per-student shuffle toggles."""
    teacher = require_admin(request)
    exam_id = request.query_params.get("exam_id")
    config = _load_exam_config(teacher["id"], exam_id=exam_id)
    sq, so = config.get("shuffle_questions", True), config.get("shuffle_options", True)
    return {"shuffle_questions": sq, "shuffle_options": so}


# ─── 42. SET SHUFFLE CONFIG ────────────────────────#

@router.post("/api/v1/admin/shuffle-config")
def admin_set_shuffle(request: Request, body: dict = Body(...)):
    """Toggle per-student question / option shuffling."""
    teacher = require_admin(request)
    tid = teacher["id"]
    exam_id = body.get("exam_id")
    fields: dict = {}
    if "shuffle_questions" in body:
        fields["shuffle_questions"] = bool(body["shuffle_questions"])
    if "shuffle_options" in body:
        fields["shuffle_options"] = bool(body["shuffle_options"])
    if not fields:
        raise HTTPException(status_code=400, detail="No shuffle fields provided")
    if tid and exam_id:
        supabase.table("exam_config").update(fields)\
            .eq("teacher_id", tid).eq("exam_id", exam_id).execute()
    else:
        update = {**({"teacher_id": tid} if tid else {"id": 1}), **fields}
        supabase.table("exam_config").upsert(update).execute()
    if _cache:
        _cache.delete(f"exam_config:{tid}:{exam_id or '_'}")
    return {
        "status": "updated",
        "shuffle_questions": fields.get("shuffle_questions"),
        "shuffle_options":   fields.get("shuffle_options"),
    }


# ─── 43. ADMIN FORCE-SUBMIT ────────────────────────#

@router.post("/api/v1/admin-submit/{session_id}")
def admin_submit(session_id: str, request: Request):
    """Force-submit a session that failed to submit properly."""
    teacher = require_admin(request)
    tid = teacher["id"]

    existing_session = _assert_session_owned(session_id, tid)
    if existing_session.get("status") == SessionStatus.COMPLETED:
        return {"status": "already_submitted"}

    from ..dependencies import _recalculate_score

    ev_result = supabase.table("violations")\
        .select("*")\
        .eq("session_key", session_id)\
        .eq("teacher_id", str(tid))\
        .order("created_at").execute()
    events = ev_result.data or []
    if not events:
        raise HTTPException(status_code=404, detail="Session not found")

    roll_number = existing_session.get("roll_number") or session_id.rsplit("_", 1)[0]
    full_name   = existing_session.get("full_name") or "Unknown"
    email       = existing_session.get("email") or "unknown@exam.com"
    for e in events:
        if e["violation_type"] == "enrollment_started" and e.get("details"):
            try:
                parts = e["details"].replace("Student: ", "")
                if "(" in parts:
                    full_name   = parts.split("(")[0].strip()
                    roll_number = parts.split("(")[1].replace(")", "").strip()
            except Exception:
                pass

    try:
        s_result = supabase.table("students").select("*")\
            .eq("roll_number", roll_number)\
            .eq("teacher_id", str(tid))\
            .execute()
        if s_result.data:
            full_name = s_result.data[0].get("full_name", full_name)
            email     = s_result.data[0].get("email", email)
    except Exception:
        pass

    answers_map: dict = {}
    for e in events:
        if e["violation_type"] == "answer_selected" and e.get("details"):
            try:
                parts = {}
                for segment in e["details"].split("|"):
                    k, _, v = segment.partition(":")
                    parts[k.strip()] = v.strip()
                if "q" in parts and "a" in parts:
                    answers_map[parts["q"]] = parts["a"]
            except Exception:
                pass

    existing_eid = existing_session.get("exam_id")
    score, total = _recalculate_score(session_id, answers_map, tid, exam_id=existing_eid)

    pct        = round((score / max(total, 1)) * 100, 1)
    now        = now_ist()
    violations = [e for e in events
                  if e["severity"] in ("high", "medium")
                  and True]
    risk = compute_risk_score(session_id, teacher_id=tid)

    sess_row = {
        "session_key":     session_id,
        "teacher_id":      str(tid),
        "roll_number":     roll_number,
        "full_name":       full_name,
        "email":           email,
        "score":           score,
        "total":           total,
        "percentage":      pct,
        "time_taken_secs": 0,
        "status":          SessionStatus.COMPLETED,
        "submitted_at":    now.isoformat(),
        "risk_score":      risk["risk_score"],
    }
    if existing_eid:
        sess_row["exam_id"] = existing_eid
    supabase.table("exam_sessions").upsert(sess_row).execute()

    if answers_map:
        ans_rows = []
        for qid, ans in answers_map.items():
            row = {"session_key": session_id, "teacher_id": str(tid),
                   "question_id": qid, "answer": ans}
            if existing_eid:
                row["exam_id"] = existing_eid
            ans_rows.append(row)
        supabase.table("answers").upsert(ans_rows).execute()

    supabase.table("violations").insert({
        "session_key":    session_id,
        "teacher_id":     str(tid),
        "violation_type": "exam_submitted",
        "severity":       "low",
        "details":        f"Admin force-submitted | Violations:{len(violations)} | Risk:{risk['risk_score']}/100",
    }).execute()

    print(f"[ForceSubmit] {session_id} score:{score}/{total} risk:{risk['risk_score']}/100")
    return {
        "status":          SessionStatus.FORCE_SUBMITTED,
        "session_id":      session_id,
        "score":           score,
        "total":           total,
        "violation_count": len(violations),
        "risk_score":      risk["risk_score"],
        "risk_label":      risk["label"],
    }
# ─── 44. REQUEST RECALIBRATION ─────────────────────────#

@router.post("/api/v1/admin/sessions/{session_id:path}/request-recalibration")
async def request_recalibration(session_id: str, request: Request):
    """End a live session and ask the student to restart for fresh calibration."""
    teacher = require_admin(request)
    tid = teacher["id"]

    sess = _assert_session_owned(session_id, tid)
    status = (sess.get("status") or "").lower()
    if status in (SessionStatus.COMPLETED, SessionStatus.SUBMITTED, SessionStatus.FORCE_SUBMITTED):
        raise HTTPException(status_code=409,
            detail="Session is already submitted; recalibration not applicable.")

    if status != SessionStatus.ABANDONED:
        try:
            supabase.table("exam_sessions")\
             .update({"status": SessionStatus.ABANDONED})\
             .eq("session_key", session_id)\
             .eq("teacher_id", str(tid)).execute()
        except Exception as e:
            print(f"[recalibration] status update failed sid={session_id}: {e}", flush=True)

    msg = ("Your teacher has requested re-calibration. Please close "
           "this exam window and re-launch from the lobby — your "
           "answers so far have been saved, but calibration will run "
           "again to recheck your gaze setup.")
    try:
        from ..routers.chat import chat_hub
        await chat_hub.teacher_send(str(tid), session_id, msg)
    except Exception as e:
        print(f"[recalibration] chat notify failed sid={session_id}: {e}", flush=True)

    try:
        viol_row = {
            "session_key":    session_id,
            "violation_type": "recalibration_requested",
            "severity":       "low",
            "details":        f"Teacher requested re-calibration. Session marked abandoned.",
            "teacher_id":     str(tid),
        }
        await _atable("violations").insert(viol_row).execute()
    except Exception as e:
        print(f"[recalibration] audit log failed sid={session_id}: {e}", flush=True)

    if _cache:
        try:
            _cache.delete(f"cal_quality:{session_id}")
        except Exception:
            pass

    return {"ok": True, "session_id": session_id, "status": "recalibration_requested"}


# ─── 45. LIVE VIEW START ─────────────────────────#

@router.post("/api/v1/admin/sessions/{session_id:path}/live-view/start")
def live_view_start(session_id: str, request: Request):
    teacher = require_admin(request)
    tid = str(teacher["id"])
    _assert_session_owned(session_id, tid)
    if _cache:
        _cache.set(f"liveview:{session_id}", {"tid": tid, "started_at": now_ist().isoformat()},
                   ttl=60)
    return {"ok": True, "session_id": session_id, "ttl_sec": 60}


# ─── 46. LIVE RISK TRIAGE ─────────────────────────#

@router.get("/api/v1/admin/sessions/{session_id:path}/triage")
def live_risk_triage_endpoint(session_id: str, request: Request):
    """One-line LLM TL;DR of a live session's recent violations."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    _assert_session_owned(session_id, tid)

    cache_key = f"triage:{session_id}"
    if _cache:
        cached = _cache.get(cache_key)
        if isinstance(cached, dict) and cached.get("summary"):
            return {**cached, "cached": True}

    try:
        sess = (supabase.table("exam_sessions").select(
                "session_key,roll_number,full_name,exam_id,started_at,current_question")
                .eq("session_key", session_id).eq("teacher_id", tid)
                .limit(1).execute()).data or []
    except Exception as e:
        print(f"[triage] session lookup failed sid={session_id}: {e}", flush=True)
        sess = []
    sess_row = sess[0] if sess else {}

    exam_id = sess_row.get("exam_id")
    exam_title = exam_id or "Exam"
    try:
        cfg = _load_exam_config(teacher_id=tid, exam_id=exam_id) if exam_id else None
        if cfg:
            exam_title = cfg.get("exam_title") or cfg.get("title") or exam_title
    except Exception:
        pass

    elapsed_minutes = None
    started_at = sess_row.get("started_at")
    if started_at:
        try:
            t0 = datetime.fromisoformat(str(started_at).replace("Z", "+00:00"))
            elapsed_minutes = max(0, int((datetime.now(timezone.utc) - t0).total_seconds() // 60))
        except Exception:
            pass

    session_meta = {
        "roll_number": sess_row.get("roll_number"),
        "full_name": sess_row.get("full_name"),
        "exam_title": exam_title,
        "elapsed_minutes": elapsed_minutes,
        "current_question": sess_row.get("current_question"),
    }

    try:
        viol_rows = (supabase.table("violations").select("*")
                     .eq("session_key", session_id).eq("teacher_id", tid)
                     .order("created_at", desc=True).limit(80)
                     .execute()).data or []
    except Exception as e:
        print(f"[triage] violation lookup failed sid={session_id}: {e}", flush=True)
        viol_rows = []

    from llm import live_risk_triage as _triage
    summary = _triage(session_meta, viol_rows)

    payload = {
        "summary": summary,
        "generated_at": now_ist().isoformat(),
        "violation_count": len(viol_rows),
    }
    if _cache and summary:
        _cache.set(cache_key, payload, ttl=60)
    return {**payload, "cached": False}


# ─── 47. LIVE VIEW KEEPALIVE ────────────────────────#

@router.post("/api/v1/admin/sessions/{session_id:path}/live-view/keepalive")
def live_view_keepalive(session_id: str, request: Request):
    teacher = require_admin(request)
    tid = str(teacher["id"])
    _assert_session_owned(session_id, tid)
    if _cache:
        _cache.set(f"liveview:{session_id}", {"tid": tid, "renewed_at": now_ist().isoformat()},
                   ttl=60)
    return {"ok": True}


# ─── 48. LIVE VIEW STOP ─────────────────────────#

@router.post("/api/v1/admin/sessions/{session_id:path}/live-view/stop")
def live_view_stop(session_id: str, request: Request):
    teacher = require_admin(request)
    tid = str(teacher["id"])
    _assert_session_owned(session_id, tid)
    if _cache:
        _cache.delete(f"liveview:{session_id}")
        _cache.delete(f"liveframe:{session_id}")
    return {"ok": True}


# ─── 49. LIVE FRAME ─────────────────────────#

@router.get("/api/v1/admin/sessions/{session_id:path}/live-frame")
def live_view_frame(session_id: str, request: Request):
    """Return the latest webcam frame for this session.

    Supports both legacy base64 (from HTTP POST) and raw bytes
    (from WebSocket binary stream).
    """
    teacher = require_admin(request)
    tid = str(teacher["id"])
    _assert_session_owned(session_id, tid)
    from starlette.responses import Response
    if not _cache:
        return Response(status_code=204)
    payload = _cache.get(f"liveframe:{session_id}")
    if not payload or not isinstance(payload, dict):
        return Response(status_code=204)

    # New WS path stores raw bytes directly
    jpeg = payload.get("jpeg_bytes")
    if jpeg:
        return Response(content=jpeg, media_type="image/jpeg",
                        headers={"Cache-Control": "no-store, max-age=0"})

    # Legacy path: base64-encoded from old HTTP POST
    b64 = payload.get("jpeg_b64")
    if not b64:
        return Response(status_code=204)
    try:
        jpeg = base64.b64decode(b64)
    except Exception:
        return Response(status_code=204)
    return Response(content=jpeg, media_type="image/jpeg",
                    headers={"Cache-Control": "no-store, max-age=0"})


# ─── 50. LIVE VIEW FORCE STOP ────────────────────────#

@router.post("/api/v1/admin/sessions/{session_id:path}/live-view/force-stop")
def live_view_force_stop(session_id: str, request: Request):
    teacher = require_admin(request)
    tid = str(teacher["id"])
    _assert_session_owned(session_id, tid)
    if _cache:
        _cache.delete(f"liveview:{session_id}")
        _cache.delete(f"liveframe:{session_id}")
    return {"ok": True}


# ─── 51. INVITES ─────────────────────────────────────#

class InviteRecipient(BaseModel):
    email: str
    full_name: str
    roll_number: str


class SendInvitesBody(BaseModel):
    recipients: list[InviteRecipient]
    exam_id: str
    custom_message: Optional[str] = None


@router.post("/api/v1/admin/invites/send")
def send_invites(body: SendInvitesBody, request: Request):
    """Send invites to a batch of students. Upserts on duplicate
    (teacher, email, exam) to avoid double-invites."""
    from ..emailer import send_invite_email
    teacher = require_admin(request)
    tid = str(teacher["id"])
    base_url = _get_invite_base_url()
    today = datetime.now(timezone.utc).date().isoformat()

    # Check remaining daily quota
    counter_rows = (supabase.table("invite_send_counters")
                    .select("count")
                    .eq("teacher_id", tid).eq("day", today).execute()).data
    used = counter_rows[0]["count"] if counter_rows else 0
    remaining = INVITE_DAILY_CAP - used
    if len(body.recipients) > remaining:
        raise HTTPException(
            status_code=429,
            detail=f"Daily cap exceeded. {remaining} remaining, {len(body.recipients)} requested."
        )

    # Fetch exam config for invite template
    exam_cfg = (supabase.table("exam_config")
                .select("*")
                .eq("teacher_id", tid).eq("exam_id", body.exam_id).execute()).data
    exam_title = exam_cfg[0].get("exam_title", body.exam_id) if exam_cfg else body.exam_id

    results = {"sent": 0, "failed": 0, "skipped": 0}
    for rec in body.recipients:
        token = _new_invite_token()
        invite_url = f"{base_url}/invite/{token}"
        download_url = f"{base_url}/download"

        # Upsert invite row
        invite_row = {
            "id": _uuid.uuid4(),
            "teacher_id": tid,
            "email": rec.email.strip().lower(),
            "full_name": rec.full_name,
            "roll_number": rec.roll_number.strip().upper(),
            "exam_id": body.exam_id,
            "token": token,
            "status": InviteStatus.SENT,
            "sent_at": now_ist().isoformat(),
            "access_code": None,
            "custom_message": body.custom_message,
        }

        # Check for existing invite to upsert
        existing = (supabase.table("student_invites")
                    .select("id,token")
                    .eq("teacher_id", tid)
                    .eq("email", rec.email.strip().lower())
                    .eq("exam_id", body.exam_id).execute()).data

        if existing:
            # Upssert: update token and mark re-sent
            (supabase.table("student_invites")
             .update({"token": token, "status": InviteStatus.SENT,
                      "sent_at": now_ist().isoformat(),
                      "custom_message": body.custom_message})
             .eq("id", existing[0]["id"]).execute())
        else:
            (supabase.table("student_invites").insert(invite_row).execute())

        # Send email
        send_result = send_invite_email(
            to_email=rec.email,
            to_name=rec.full_name,
            exam_title=exam_title,
            invite_url=invite_url,
            download_url=download_url,
            roll_number=rec.roll_number,
            teacher_name=teacher.get("email"),
        )
        if send_result.ok:
            # Stamp provider_msg_id on the invite
            (supabase.table("student_invites")
             .update({"provider_msg_id": send_result.provider_msg_id})
             .eq("teacher_id", tid).eq("email", rec.email.strip().lower())
             .eq("exam_id", body.exam_id).execute())
            results["sent"] += 1
        else:
            results["failed"] += 1

    # Bump counter
    if existing_counter := counter_rows:
        (supabase.table("invite_send_counters")
         .update({"count": used + len(body.recipients)})
         .eq("teacher_id", tid).eq("day", today).execute())
    else:
        (supabase.table("invite_send_counters")
         .insert({"teacher_id": tid, "day": today, "count": len(body.recipients)}).execute())

    return results


@router.get("/api/v1/admin/invites")
def list_invites(request: Request, exam_id: Optional[str] = None):
    """List all invites for the authenticated teacher."""
    teacher = require_admin(request)
    tid = str(teacher["id"])
    base_url = _get_invite_base_url()

    query = (supabase.table("student_invites")
             .select("*")
             .eq("teacher_id", tid)
             .order("sent_at", desc=True))
    if exam_id:
        query = query.eq("exam_id", exam_id)
    result = query.execute()

    invites = []
    for row in result.data or []:
        invites.append({
            "id": row.get("id"),
            "email": row.get("email"),
            "full_name": row.get("full_name"),
            "roll_number": row.get("roll_number"),
            "exam_id": row.get("exam_id"),
            "token": row.get("token"),
            "status": row.get("status"),
            "invite_url": f"{base_url}/invite/{row.get('token', '')}",
            "sent_at": row.get("sent_at"),
            "opened_at": row.get("opened_at"),
            "bounced_at": row.get("bounced_at"),
            "provider_msg_id": row.get("provider_msg_id"),
        })
    return {"invites": invites, "total": len(invites)}


@router.delete("/api/v1/admin/invites/{invite_id}")
def revoke_invite(invite_id: str, request: Request):
    """Revoke a single invite by ID."""
    teacher = require_admin(request)
    tid = str(teacher["id"])

    result = (supabase.table("student_invites")
              .select("id,teacher_id,status")
              .eq("id", invite_id).execute())
    if not result.data:
        raise HTTPException(status_code=404, detail="Invite not found")
    if result.data[0].get("teacher_id") != tid:
        raise HTTPException(status_code=403, detail="Not your invite")

    (supabase.table("student_invites")
     .update({"status": InviteStatus.REVOKED, "revoked_at": now_ist().isoformat()})
     .eq("id", invite_id).execute())
    return {"ok": True, "invite_id": invite_id}
